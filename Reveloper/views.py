from io import BytesIO
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login as auth_login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.utils import timezone
from django.utils.html import strip_tags
from django.http import HttpResponse
from django.conf import settings
from django.db.models import Q
from .models import TareaPorDesarrollar, Proyecto, Usuario, Evaluacion, EvaluacionConfig
from .forms import TareaPorDesarrollarForm, EvaluacionForm
import io
import json
import base64
import openpyxl
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Image, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from matplotlib.colors import ListedColormap
import matplotlib.pyplot as plt
import os
import tempfile
import matplotlib
matplotlib.use('Agg')


# Función de Verificación para Administradores


def es_admin(user):
    return user.is_superuser

# Vista para el inicio de sesión personalizado


def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            auth_login(request, user)
            return redirect('/reveloper/home/')
    else:
        form = AuthenticationForm()
    return render(request, 'registration/login.html', {'form': form})

# Vista para cerrar sesión


def logout_view(request):
    logout(request)
    return redirect('login')

# Vista protegida para el home, accesible solo para usuarios autenticados


@login_required
def home(request):
    tareas_pendientes = TareaPorDesarrollar.objects.filter(
        usuario=request.user, estado='pendiente')
    context = {
        'usuario': request.user,
        # Serializar a JSON
        'tareas_pendientes': json.dumps(list(tareas_pendientes.values('titulo')))
    }
    return render(request, 'home.html', context)


# Vista para imprimir las carpetas de plantillas configuradas


def print_template_dirs(request):
    print(settings.TEMPLATES[0]['DIRS'])
    return HttpResponse('Template Dirs Printed')

# Vista para generar los datos que se mostrarán en el dashboard


@login_required
@user_passes_test(es_admin)
def dashboard(request):
    usuarios = Usuario.objects.all()
    labels = [usuario.username for usuario in usuarios]
    data = [usuario.tareas_completadas for usuario in usuarios]

    # Datos para el gráfico de distribución de tareas pendientes por usuario
    tareas_usuario_data = {usuario.username: {
        'pendiente': TareaPorDesarrollar.objects.filter(usuario=usuario, estado='pendiente').count()
    } for usuario in usuarios}

    context = {
        'usuarios': usuarios,
        'labels_json': json.dumps(labels),
        'data_json': json.dumps(data),
        'tareas_usuario_labels_json': json.dumps(labels),
        'tareas_usuario_data_json': json.dumps(tareas_usuario_data)
    }
    return render(request, 'admin/dashboard.html', context)


# Vista para la página de usuarios, accesible solo para usuarios autenticados


@login_required
@user_passes_test(es_admin)
def usuarios(request):
    usuarios = Usuario.objects.all()
    for usuario in usuarios:
        usuario.tareas_asignadas = TareaPorDesarrollar.objects.filter(
            usuario=usuario)

    context = {
        'usuarios': usuarios
    }
    return render(request, 'usuarios.html', context)


# Vista para la página de proyectos, accesible solo para usuarios autenticados


@login_required
def proyectos(request):
    proyectos = Proyecto.objects.all()
    for proyecto in proyectos:
        proyecto.tareas = TareaPorDesarrollar.objects.filter(proyecto=proyecto)

    context = {
        'proyectos': proyectos
    }

    return render(request, 'proyectos.html', context)

# Vista para la página de evaluaciones, accesible solo para usuarios autenticados


@login_required
def evaluaciones(request):
    if request.user.is_superuser:  # El administrador puede ver todas las evaluaciones
        evaluaciones = Evaluacion.objects.all()
    else:  # Otros usuarios solo ven sus evaluaciones
        evaluaciones = Evaluacion.objects.filter(usuario=request.user)

    return render(request, 'evaluaciones.html', {'evaluaciones': evaluaciones})

# Vista para la página de tareas, accesible solo para usuarios autenticados


@login_required
def tareas_por_desarrollar(request):
    if request.user.is_superuser:
        # Si es administrador, mostrar todas las tareas
        tareas = TareaPorDesarrollar.objects.select_related(
            'proyecto', 'usuario').all()
    else:
        # Si no es administrador, mostrar solo las tareas asignadas al usuario
        tareas = TareaPorDesarrollar.objects.select_related(
            'proyecto', 'usuario').filter(usuario=request.user)
    return render(request, 'tareas.html', {'tareas': tareas})


# Vista para Marcar Tarea como "En Revisión"


@login_required
def marcar_tarea_en_revision(request, tarea_id):
    tarea = get_object_or_404(TareaPorDesarrollar, id=tarea_id)
    if tarea.usuario == request.user:
        if tarea.estado == 'pendiente':
            tarea.estado = 'en progreso'
        elif tarea.estado == 'en progreso':
            tarea.estado = 'en revision'
        tarea.save()
    return redirect('tareas_por_desarrollar')


# Vista para Revisar Tareas y Crear Evaluaciones

@login_required
@user_passes_test(es_admin)
def revisar_tareas(request):
    tareas_en_revision = TareaPorDesarrollar.objects.filter(
        estado='en revision')

    # Obtener la nota máxima de la configuración de evaluación
    evaluacion_config = EvaluacionConfig.objects.first()
    # Valor por defecto
    nota_maxima = evaluacion_config.nota_maxima if evaluacion_config else 100

    if request.method == 'POST':
        tarea_id = request.POST.get('tarea_id')
        tarea = get_object_or_404(TareaPorDesarrollar, id=tarea_id)

        # Procesar el formulario de evaluación
        form = EvaluacionForm(request.POST)
        if form.is_valid():
            tarea.estado = 'completada'
            tarea.save()

            # Obtener los puntajes del formulario
            tiempo_entrega = form.cleaned_data['tiempo_entrega']
            complejidad_tarea = form.cleaned_data['complejidad_tarea']
            cumplimiento_requerimientos = form.cleaned_data['cumplimiento_requerimientos']
            calidad_codigo = form.cleaned_data['calidad_codigo']

            # Calcular la calificación automática
            calificacion = tiempo_entrega + complejidad_tarea + \
                cumplimiento_requerimientos + calidad_codigo
            comentarios = f"Evaluación automática: Tiempo de Entrega: {tiempo_entrega}, Complejidad de la Tarea: {
                complejidad_tarea}, Cumplimiento de Requerimientos: {cumplimiento_requerimientos}, Calidad del Código: {calidad_codigo}"

            # Crear evaluación automática
            Evaluacion.objects.create(
                titulo=f"Evaluación para {tarea.titulo}",
                comentarios=comentarios,
                fecha_evaluacion=timezone.now(),
                proyecto=tarea.proyecto,
                usuario=tarea.usuario,
                calificacion=calificacion,
                tarea=tarea
            )

            return redirect('revisar_tareas')

    else:
        form = EvaluacionForm()

    return render(request, 'revisar_tareas.html', {'tareas': tareas_en_revision, 'form': form, 'nota_maxima': nota_maxima})


@login_required
@user_passes_test(es_admin)
def crear_tarea(request):
    if request.method == 'POST':
        form = TareaPorDesarrollarForm(request.POST)
        if form.is_valid():
            tarea = form.save(commit=False)
            tarea.usuario = request.user  # Asignar el usuario autenticado
            tarea.save()  # Guardar la tarea con el usuario asignado
            return redirect('proyectos')
    else:
        form = TareaPorDesarrollarForm()
    return render(request, 'crear_tarea.html', {'form': form})


@login_required
@user_passes_test(es_admin)
def editar_tarea(request, tarea_id):
    tarea = get_object_or_404(TareaPorDesarrollar, pk=tarea_id)
    if request.method == 'POST':
        form = TareaPorDesarrollarForm(request.POST, instance=tarea)
        if form.is_valid():
            form.save()
            # Redirige a la lista de proyectos o donde prefieras
            return redirect('proyectos')
        else:
            form = TareaPorDesarrollarForm(instance=tarea)
        return render(request, 'editar_tarea.html', {'form': form})


# Vista para la búsqueda
@login_required
@user_passes_test(es_admin)
def busqueda(request):
    proyectos = Proyecto.objects.all()
    tareas = TareaPorDesarrollar.objects.all()
    context = {
        'proyectos': proyectos,
        'tareas': tareas,
    }
    return render(request, 'busqueda.html', context)


@login_required
@user_passes_test(es_admin)
def buscar_usuarios(request):
    nombre_o_apellido = request.GET.get('nombre_o_apellido', '')
    fecha_alta_desde = request.GET.get('fecha_alta_desde', '')
    fecha_alta_hasta = request.GET.get('fecha_alta_hasta', '')
    resultados_usuarios = []

    query = Q()
    if nombre_o_apellido:
        query &= Q(first_name__icontains=nombre_o_apellido) | Q(
            last_name__icontains=nombre_o_apellido)
    if fecha_alta_desde and fecha_alta_hasta:
        query &= Q(date_joined__gte=fecha_alta_desde) & Q(
            date_joined__lte=fecha_alta_hasta)

    if query:
        resultados_usuarios = Usuario.objects.filter(query)

    request.session['resultados_usuarios'] = [
        usuario.id for usuario in resultados_usuarios]

    context = {
        'resultados_usuarios': resultados_usuarios,
        'usuarios': Usuario.objects.all()
    }

    return render(request, 'busqueda.html', context)


@login_required
@user_passes_test(es_admin)
def buscar_proyectos(request):
    fecha_inicio_desde = request.GET.get('fecha_inicio_desde')
    fecha_inicio_hasta = request.GET.get('fecha_inicio_hasta')
    proyecto_id = request.GET.get('proyecto_id')
    titulo_palabras = request.GET.get('titulo_palabras')
    resultados = []

    query = Q()
    if fecha_inicio_desde and fecha_inicio_hasta:
        fecha_inicio_desde = timezone.make_aware(
            datetime.strptime(fecha_inicio_desde, '%Y-%m-%d'))
        fecha_inicio_hasta = timezone.make_aware(
            datetime.strptime(fecha_inicio_hasta, '%Y-%m-%d'))
        query &= Q(fecha_inicio__gte=fecha_inicio_desde) & Q(
            fecha_inicio__lte=fecha_inicio_hasta)
    if proyecto_id:
        query &= Q(id=proyecto_id)
    if titulo_palabras:
        query &= Q(nombre__icontains=titulo_palabras)

    if query:
        resultados = Proyecto.objects.filter(query)
        for proyecto in resultados:
            proyecto.tareas = TareaPorDesarrollar.objects.filter(
                proyecto=proyecto)

    request.session['resultados'] = [proyecto.id for proyecto in resultados]

    context = {
        'resultados': resultados,
        'proyectos': Proyecto.objects.all()
    }

    return render(request, 'busqueda.html', context)


@login_required
@user_passes_test(es_admin)
def buscar_tareas(request):
    fecha_inicio_desde_tarea = request.GET.get('fecha_inicio_desde_tarea')
    fecha_inicio_hasta_tarea = request.GET.get('fecha_inicio_hasta_tarea')
    tarea_id = request.GET.get('tarea_id')
    titulo_palabras = request.GET.get('titulo_palabras')
    resultados_tareas = []

    query = Q()
    if fecha_inicio_desde_tarea and fecha_inicio_hasta_tarea:
        fecha_inicio_desde_tarea = timezone.make_aware(
            datetime.strptime(fecha_inicio_desde_tarea, '%Y-%m-%d'))
        fecha_inicio_hasta_tarea = timezone.make_aware(
            datetime.strptime(fecha_inicio_hasta_tarea, '%Y-%m-%d'))
        query &= Q(fecha_creacion__gte=fecha_inicio_desde_tarea) & Q(
            fecha_creacion__lte=fecha_inicio_hasta_tarea)
    if tarea_id:
        query &= Q(id=tarea_id)
    if titulo_palabras:
        query &= Q(titulo__icontains=titulo_palabras)

    if query:
        resultados_tareas = TareaPorDesarrollar.objects.filter(query)

    request.session['resultados_tareas'] = [
        tarea.id for tarea in resultados_tareas]

    context = {
        'resultados_tareas': resultados_tareas,
        'tareas': TareaPorDesarrollar.objects.all()
    }

    return render(request, 'busqueda.html', context)


@login_required
def generate_pdf(request):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Título del documento
    logo_path = "Reveloper/static/img/logos/logo-reveloper.png"
    title = "Informe de Proyectos"
    username = f"Usuario: {request.user.username}"
    subtitle = "Lista de Proyectos:"

    # Estilos de párrafo
    styleN = styles["BodyText"]
    styleH = styles["Heading1"]
    styleN = ParagraphStyle('Normal', spaceAfter=5, leading=10)
    styleH = ParagraphStyle('Heading1', spaceAfter=10, fontSize=16,
                            textColor=colors.darkblue, fontName='Helvetica-Bold')

    # Estilo para el ID y nombre del proyecto
    styleID = ParagraphStyle('Normal', spaceAfter=10,
                             fontSize=14, textColor=colors.black, leading=15)

    # Estilo para el nombre de las tareas
    styleTaskTitle = ParagraphStyle('Normal', spaceAfter=8, fontSize=12,
                                    textColor=colors.black, leading=12, fontName='Helvetica-Bold')

    # Estilo para el desarrollador asignado
    styleAssignedTo = ParagraphStyle(
        'Normal', spaceAfter=8, fontSize=12, textColor=colors.black, leading=12)

    # Estilo para el nombre del usuario
    styleUsername = ParagraphStyle(
        'Normal', spaceAfter=12, fontSize=14, textColor=colors.black, fontName='Helvetica-Bold')

    # Añadir logotipo y títulos
    story.append(Paragraph(
        f'<img src="{logo_path}" width="100" height="50" valign="middle"/>', styleN))
    story.append(Spacer(1, 20))
    story.append(Paragraph(title, styleH))
    story.append(Spacer(1, 12))
    story.append(Paragraph(username, styleUsername))
    story.append(Spacer(1, 12))
    story.append(Paragraph(subtitle, styleN))
    story.append(Spacer(1, 12))

    # Obtener datos del contexto
    projects = Proyecto.objects.all()
    for project in projects:
        story.append(Paragraph(f"ID: {project.id}, Nombre: {
                     project.nombre}", styleID))
        story.append(Spacer(1, 10))

        # Ajustar texto de la descripción
        story.append(Paragraph(f"Descripción: {project.descripcion}", styleN))
        story.append(Paragraph(f"Fecha de Inicio: {
                     project.fecha_inicio}", styleN))
        story.append(Paragraph(f"Fecha de Fin: {project.fecha_fin}", styleN))
        story.append(
            Paragraph(f"Estado: {project.get_estado_display()}", styleN))
        story.append(Spacer(1, 10))

        # Añadir tareas al PDF
        tareas = TareaPorDesarrollar.objects.filter(
            proyecto=project).select_related('usuario')
        for tarea in tareas:
            story.append(Paragraph(f"Tarea: {tarea.titulo}", styleTaskTitle))
            story.append(Paragraph(f"Asignado a: {tarea.usuario.first_name} {
                         tarea.usuario.last_name}", styleAssignedTo))
            story.append(Paragraph(f"Estado: {tarea.estado}", styleN))
            story.append(Paragraph(f"Fecha de Creación: {
                         tarea.fecha_creacion}", styleN))
            story.append(Paragraph(f"Fecha de Vencimiento: {
                         tarea.fecha_vencimiento}", styleN))
            story.append(Spacer(1, 10))

        story.append(Spacer(1, 12))
        story.append(HRFlowable(width="100%", thickness=1,
                     color=colors.grey, spaceBefore=1, spaceAfter=24))

    doc.build(story)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


@login_required
def generate_task_pdf(request):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Título del documento
    logo_path = "Reveloper/static/img/logos/logo-reveloper.png"
    title = "Informe de Tareas"
    username = f"Usuario: {request.user.username}"

    # Estilos de párrafo
    styleN = styles["BodyText"]
    styleH = styles["Heading1"]
    styleN = ParagraphStyle('Normal', spaceAfter=5, leading=10)
    styleH = ParagraphStyle('Heading1', spaceAfter=10, fontSize=16,
                            textColor=colors.darkblue, fontName='Helvetica-Bold')

    # Estilo para el título de las tareas
    styleTaskTitle = ParagraphStyle('Normal', spaceAfter=8, fontSize=12,
                                    textColor=colors.black, leading=12, fontName='Helvetica-Bold')

    # Estilo para el asignado a
    styleAssignedTo = ParagraphStyle(
        'Normal', spaceAfter=8, fontSize=12, textColor=colors.black, leading=12, fontName='Helvetica-Bold')

    # Estilo para el nombre del usuario
    styleUsername = ParagraphStyle(
        'Normal', spaceAfter=12, fontSize=14, textColor=colors.black, fontName='Helvetica-Bold')

    # Añadir logotipo y títulos
    story.append(Paragraph(
        f'<img src="{logo_path}" width="100" height="50" valign="middle"/>', styleN))
    # Aumentar el espaciado entre el logo y el título
    story.append(Spacer(1, 20))
    story.append(Paragraph(title, styleH))
    story.append(Spacer(1, 12))
    story.append(Paragraph(username, styleUsername))
    story.append(Spacer(1, 12))

    # Verificar si el usuario es administrador
    if request.user.is_superuser:
        tareas = TareaPorDesarrollar.objects.all()
        subtitle = "Lista de Todas las Tareas:"
    else:
        tareas = TareaPorDesarrollar.objects.filter(usuario=request.user)
        subtitle = "Lista de Tareas Asignadas:"

    story.append(Paragraph(subtitle, styleN))
    story.append(Spacer(1, 12))

    for tarea in tareas:
        story.append(Paragraph(f"Título: {tarea.titulo}", styleTaskTitle))
        story.append(Paragraph(f"Descripción: {tarea.descripcion}", styleN))
        story.append(Paragraph(f"Estado: {tarea.estado}", styleN))
        story.append(Paragraph(f"Fecha de Creación: {
                     tarea.fecha_creacion}", styleN))
        story.append(Paragraph(f"Fecha de Vencimiento: {
                     tarea.fecha_vencimiento}", styleN))
        story.append(Paragraph(f"Proyecto: {tarea.proyecto.nombre}", styleN))
        story.append(Paragraph(f"Asignado a: {tarea.usuario.first_name} {
                     tarea.usuario.last_name}", styleAssignedTo))
        story.append(Spacer(1, 10))

        # Añadir una línea horizontal de margen a margen para separar tareas
        story.append(HRFlowable(width="100%", thickness=1,
                     color=colors.grey, spaceBefore=1, spaceAfter=24))

    # Cerrar y guardar el PDF
    doc.build(story)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


@login_required
def generate_evaluation_pdf(request):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Título del documento
    logo_path = "Reveloper/static/img/logos/logo-reveloper.png"
    title = "Informe de Evaluaciones"
    username = f"Usuario: {request.user.username}"

    # Estilos de párrafo
    styleN = styles["BodyText"]
    styleH = styles["Heading1"]
    styleN = ParagraphStyle('Normal', spaceAfter=5, leading=10)
    styleH = ParagraphStyle('Heading1', spaceAfter=10, fontSize=16,
                            textColor=colors.darkblue, fontName='Helvetica-Bold')

    # Estilo para el título de las evaluaciones
    styleEvaluationTitle = ParagraphStyle(
        'Normal', spaceAfter=8, fontSize=12, textColor=colors.black, leading=12, fontName='Helvetica-Bold')

    # Estilo para la calificación
    styleCalificacion = ParagraphStyle(
        'Normal', spaceAfter=8, fontSize=14, textColor=colors.black, leading=12, fontName='Helvetica-Bold')

    # Estilo para el asignado a
    styleAssignedTo = ParagraphStyle(
        'Normal', spaceAfter=8, fontSize=12, textColor=colors.black, leading=12)

    # Estilo para el nombre del usuario
    styleUsername = ParagraphStyle(
        'Normal', spaceAfter=12, fontSize=14, textColor=colors.black, fontName='Helvetica-Bold')

    # Añadir logotipo y títulos
    story.append(Paragraph(
        f'<img src="{logo_path}" width="100" height="50" valign="middle"/>', styleN))
    # Aumentar el espaciado entre el logo y el título
    story.append(Spacer(1, 20))
    story.append(Paragraph(title, styleH))
    story.append(Spacer(1, 12))
    story.append(Paragraph(username, styleUsername))
    story.append(Spacer(1, 12))

    # Verificar si el usuario es administrador
    if request.user.is_superuser:
        evaluaciones = Evaluacion.objects.all()
        subtitle = "Lista de Todas las Evaluaciones:"
    else:
        evaluaciones = Evaluacion.objects.filter(usuario=request.user)
        subtitle = "Lista de Evaluaciones Asignadas:"

    story.append(Paragraph(subtitle, styleN))
    story.append(Spacer(1, 12))

    for evaluacion in evaluaciones:
        story.append(
            Paragraph(f"Título: {evaluacion.titulo}", styleEvaluationTitle))
        story.append(Paragraph(f"Calificación: {
                     evaluacion.calificacion}", styleCalificacion))
        if evaluacion.tarea:
            story.append(
                Paragraph(f"Tarea: {evaluacion.tarea.titulo}", styleN))
        else:
            story.append(Paragraph("Tarea: Sin Tarea Asignada", styleN))
        story.append(Paragraph(f"Asignado a: {evaluacion.usuario.first_name} {
                     evaluacion.usuario.last_name}", styleAssignedTo))
        story.append(Paragraph(f"Fecha de Evaluación: {
                     evaluacion.fecha_evaluacion}", styleN))
        story.append(
            Paragraph(f"Proyecto: {evaluacion.proyecto.nombre}", styleN))
        story.append(Spacer(1, 10))

        # Añadir una línea horizontal de margen a margen para separar evaluaciones
        story.append(HRFlowable(width="100%", thickness=1,
                     color=colors.grey, spaceBefore=1, spaceAfter=24))

    # Cerrar y guardar el PDF
    doc.build(story)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


@login_required
def generate_user_pdf(request):
    # Esta vista sólo debe ser accesible para administradores
    if not request.user.is_superuser:
        return HttpResponse("No tienes permiso para acceder a esta página.", status=403)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Título del documento
    logo_path = "Reveloper/static/img/logos/logo-reveloper.png"
    title = "Informe de Usuarios"
    generated_by = f"Generado por: {request.user.username}"
    subtitle = "Lista de Usuarios:"

    # Estilos de párrafo
    styleN = styles["BodyText"]
    styleH = styles["Heading1"]
    styleN = ParagraphStyle('Normal', spaceAfter=5, leading=10)
    styleH = ParagraphStyle('Heading1', spaceAfter=10, fontSize=16,
                            textColor=colors.darkblue, fontName='Helvetica-Bold')

    # Estilo para el nombre de los usuarios
    styleUserName = ParagraphStyle('Normal', spaceAfter=10, fontSize=14,
                                   textColor=colors.black, leading=15, fontName='Helvetica-Bold')

    # Añadir logotipo y títulos
    story.append(Paragraph(
        f'<img src="{logo_path}" width="100" height="50" valign="middle"/>', styleN))
    # Aumentar el espaciado entre el logo y el título
    story.append(Spacer(1, 20))
    story.append(Paragraph(title, styleH))
    story.append(Spacer(1, 12))
    story.append(Paragraph(generated_by, styleUserName))
    story.append(Spacer(1, 12))
    story.append(Paragraph(subtitle, styleN))
    story.append(Spacer(1, 12))

    usuarios = Usuario.objects.all()
    for usuario in usuarios:
        story.append(Paragraph(f"Nombre: {usuario.first_name} {
                     usuario.last_name}", styleUserName))
        story.append(Paragraph(f"Username: {usuario.username}", styleN))
        story.append(Paragraph(f"Email: {usuario.email}", styleN))
        story.append(Paragraph(f"Fecha de Registro: {
                     usuario.date_joined}", styleN))

        # Añadir tareas asignadas
        tareas_asignadas = TareaPorDesarrollar.objects.filter(usuario=usuario)
        if tareas_asignadas:
            story.append(Spacer(1, 10))
            story.append(Paragraph("Tareas Asignadas:", styleUserName))
            for tarea in tareas_asignadas:
                story.append(Paragraph(f"- Tarea: {tarea.titulo}", styleN))
                story.append(Paragraph(f"  Estado: {tarea.estado}", styleN))
                story.append(Paragraph(f"  Fecha de Creación: {
                             tarea.fecha_creacion}", styleN))
                story.append(Paragraph(f"  Fecha de Vencimiento: {
                             tarea.fecha_vencimiento}", styleN))
                story.append(Spacer(1, 5))

        # Añadir una línea horizontal de margen a margen para separar usuarios
        story.append(HRFlowable(width="100%", thickness=1,
                     color=colors.grey, spaceBefore=1, spaceAfter=24))

    # Cerrar y guardar el PDF
    doc.build(story)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


@login_required
def vista_evaluaciones(request):
    if request.user.is_superuser:  # Si es administrador, muestra todas las evaluaciones
        evaluaciones = Evaluacion.objects.all()
    else:  # Si es usuario normal, muestra solo sus evaluaciones
        evaluaciones = Evaluacion.objects.filter(usuario=request.user)

    graph = generar_grafico_evaluaciones(request)
    context = {
        'evaluaciones': evaluaciones,
        'graph': graph
    }
    return render(request, 'evaluaciones.html', context)


def generar_grafico_evaluaciones(request, desarrollador):
    if request.user.is_superuser:
        evaluaciones = Evaluacion.objects.filter(usuario=desarrollador)
    else:
        evaluaciones = Evaluacion.objects.filter(usuario=request.user)

    fechas = [evaluacion.fecha_evaluacion for evaluacion in evaluaciones]
    puntajes = [evaluacion.calificacion for evaluacion in evaluaciones]

    # Definir una paleta de colores fija
    colores = ListedColormap(["#FF6347", "#4682B4", "#8A2BE2",
                             "#5F9EA0", "#7FFF00", "#FF7F50", "#FFD700", "#DC143C"]).colors

    plt.figure(figsize=(10, 5))
    # Usar los colores necesarios
    plt.bar(fechas, puntajes, color=colores[:len(fechas)])
    plt.title(f'Evaluaciones de {desarrollador.username}')
    plt.xlabel('Fecha')
    plt.ylabel('Puntaje')
    plt.xticks(rotation=45)
    plt.tight_layout()

    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    plt.close()
    buffer.seek(0)
    image_png = buffer.getvalue()
    graph = base64.b64encode(image_png).decode('utf-8')
    buffer.close()

    return graph


def generar_informe_grafico_pdf_admin(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="informe_grafico_evaluaciones.pdf"'

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    initial_height = height - 40  # Espacio inicial desde la parte superior

    p.drawString(100, initial_height, "Informe de Evaluaciones")
    height = initial_height - 20  # Ajustar después del título principal

    # Filtra los usuarios que no son administradores
    desarrolladores = Usuario.objects.filter(is_staff=False)

    for desarrollador in desarrolladores:
        graph = generar_grafico_evaluaciones(request, desarrollador)
        graph_data = base64.b64decode(graph)

        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as temp_file:
            temp_file.write(graph_data)
            temp_file_path = temp_file.name

        p.drawString(100, height, f"Desarrollador: {desarrollador.username}")
        height -= 20  # Ajustar espacio entre el nombre del desarrollador y el gráfico

        p.drawImage(temp_file_path, 50, height - 250, width=500, height=200)
        height -= 270  # Ajustar espacio después del gráfico

        if height < 320:  # Añadir una nueva página si no hay suficiente espacio
            p.showPage()
            height = initial_height

        os.remove(temp_file_path)

    p.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response


def generar_informe_grafico_pdf_desarrollador(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="informe_grafico_evaluaciones.pdf"'

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.drawString(100, height - 40, "Informe de Evaluaciones")

    desarrollador = request.user  # Asumimos que el desarrollador está autenticado
    graph = generar_grafico_evaluaciones(request, desarrollador)
    graph_data = base64.b64decode(graph)

    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as temp_file:
        temp_file.write(graph_data)
        temp_file_path = temp_file.name

    p.drawString(100, height - 60, f"Desarrollador: {desarrollador.username}")
    p.drawImage(temp_file_path, 50, height - 300, width=500, height=200)

    p.showPage()
    p.save()

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    os.remove(temp_file_path)

    return response

# INFORMES PDF BÚSQUEDA
# Informe PDF para la búsqueda de proyectos


@login_required
@user_passes_test(es_admin)
def generar_informe_pdf_busqueda(request):
    from django.utils.html import strip_tags

    fecha_inicio_desde = request.GET.get('fecha_inicio_desde', '')
    fecha_inicio_hasta = request.GET.get('fecha_inicio_hasta', '')
    proyecto_id = request.GET.get('proyecto_id', '')
    titulo_palabras = request.GET.get('titulo_palabras', '')
    resultados_ids = request.session.get('resultados', [])
    resultados = Proyecto.objects.filter(id__in=resultados_ids)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Título del documento
    logo_path = "Reveloper/static/img/logos/logo-reveloper.png"
    title = "Informe de Búsqueda de Proyectos"
    username = f"Usuario: {request.user.username}"
    subtitle = f"Lista de Proyectos iniciados entre {
        fecha_inicio_desde} y {fecha_inicio_hasta}:"

    # Estilos de párrafo
    styleN = styles["BodyText"]
    styleH = styles["Heading1"]
    styleN = ParagraphStyle('Normal', spaceAfter=5, leading=10)
    styleH = ParagraphStyle('Heading1', spaceAfter=10, fontSize=16,
                            textColor=colors.darkblue, fontName='Helvetica-Bold')

    # Estilo para el ID y nombre del proyecto
    styleID = ParagraphStyle('Normal', spaceAfter=10,
                             fontSize=14, textColor=colors.black, leading=15)

    # Estilo para el nombre de las tareas
    styleTaskTitle = ParagraphStyle('Normal', spaceAfter=8, fontSize=12,
                                    textColor=colors.black, leading=12, fontName='Helvetica-Bold')

    # Estilo para el desarrollador asignado
    styleAssignedTo = ParagraphStyle(
        'Normal', spaceAfter=8, fontSize=12, textColor=colors.black, leading=12)

    # Estilo para el nombre del usuario
    styleUsername = ParagraphStyle(
        'Normal', spaceAfter=12, fontSize=14, textColor=colors.black, fontName='Helvetica-Bold')

    # Añadir logotipo y títulos
    story.append(Paragraph(
        f'<img src="{logo_path}" width="100" height="50" valign="middle"/>', styleN))
    story.append(Spacer(1, 20))
    story.append(Paragraph(title, styleH))
    story.append(Spacer(1, 12))
    story.append(Paragraph(username, styleUsername))
    story.append(Spacer(1, 12))
    story.append(Paragraph(subtitle, styleN))
    story.append(Spacer(1, 12))

    # Agregar datos al documento
    for proyecto in resultados:
        # Strip HTML tags from project title
        project_title = strip_tags(
            f"<u>ID: {proyecto.id}, Nombre: {proyecto.nombre}</u>")
        story.append(Paragraph(project_title, styleID))
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"Descripción: {proyecto.descripcion}", styleN))
        story.append(
            Paragraph(f"Fecha de Inicio: {proyecto.fecha_inicio}", styleN))
        story.append(Paragraph(f"Fecha de Fin: {proyecto.fecha_fin}", styleN))
        story.append(
            Paragraph(f"Estado: {proyecto.get_estado_display()}", styleN))
        story.append(Spacer(1, 10))

        # Añadir tareas al PDF
        tareas = TareaPorDesarrollar.objects.filter(proyecto=proyecto)
        for tarea in tareas:
            story.append(Paragraph(f"Tarea: {tarea.titulo}", styleTaskTitle))
            story.append(Paragraph(
                f"Asignado a: {tarea.usuario.first_name} {tarea.usuario.last_name}", styleAssignedTo))
            story.append(Paragraph(f"Estado: {tarea.estado}", styleN))
            story.append(
                Paragraph(f"Fecha de Creación: {tarea.fecha_creacion}", styleN))
            story.append(
                Paragraph(f"Fecha de Vencimiento: {tarea.fecha_vencimiento}", styleN))
            story.append(Spacer(1, 10))

        story.append(Spacer(1, 12))
        story.append(HRFlowable(width="100%", thickness=1,
                     color=colors.grey, spaceBefore=1, spaceAfter=24))

    # Cerrar y guardar el PDF
    doc.build(story)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


@login_required
@user_passes_test(es_admin)
def generar_informe_pdf_tareas(request):
    tarea_id = request.GET.get('tarea_id', '')
    titulo_palabras = request.GET.get('titulo_palabras', '')
    resultados_tareas = request.session.get('resultados_tareas', [])

    # Asegurarse de que la consulta esté utilizando los IDs de las tareas
    tareas = TareaPorDesarrollar.objects.filter(id__in=resultados_tareas)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Título del documento
    # Asegúrate de que la ruta al logotipo sea correcta
    logo_path = "Reveloper/static/img/logos/logo-reveloper.png"
    title = "Informe de Búsqueda de Tareas"
    subtitle = "Lista de Tareas encontradas:"

    styleN = ParagraphStyle('Normal', spaceAfter=5, leading=10)
    styleH = ParagraphStyle('Heading1', spaceAfter=10, fontSize=16,
                            textColor=colors.darkblue, fontName='Helvetica-Bold')

    # Añadir logotipo y títulos
    logo = Image(logo_path, width=2*inch, height=1*inch)
    logo.hAlign = 'LEFT'
    story.append(logo)
    story.append(Spacer(1, 20))
    story.append(Paragraph(title, styleH))
    story.append(Spacer(1, 12))
    story.append(Paragraph(subtitle, styleN))
    story.append(Spacer(1, 12))

    for tarea in tareas:
        story.append(Paragraph(f"ID: {tarea.id}", styleN))
        story.append(Paragraph(f"Título: {tarea.titulo}", styleN))
        story.append(Paragraph(f"Descripción: {tarea.descripcion}", styleN))
        story.append(Paragraph(
            f"Asignado a: {tarea.usuario.first_name} {tarea.usuario.last_name}", styleN))
        story.append(
            Paragraph(f"Fecha de Creación: {tarea.fecha_creacion}", styleN))
        story.append(
            Paragraph(f"Fecha de Vencimiento: {tarea.fecha_vencimiento}", styleN))
        story.append(Paragraph(f"Estado: {tarea.estado}", styleN))
        story.append(Spacer(1, 12))
        story.append(HRFlowable(width="100%", thickness=1,
                     color=colors.grey, spaceBefore=1, spaceAfter=24))

    doc.build(story)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


@login_required
@user_passes_test(es_admin)
def generar_informe_pdf_usuarios(request):
    nombre_o_apellido = request.GET.get('nombre_o_apellido', '')
    fecha_alta_desde = request.GET.get('fecha_alta_desde', '')
    fecha_alta_hasta = request.GET.get('fecha_alta_hasta', '')
    resultados_usuarios = request.session.get('resultados_usuarios', [])
    usuarios = Usuario.objects.filter(id__in=resultados_usuarios)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Título del documento
    # Asegúrate de que la ruta al logotipo sea correcta
    logo_path = "Reveloper/static/img/logos/logo-reveloper.png"
    title = "Informe de Búsqueda de Usuarios"
    subtitle = "Lista de Usuarios encontrados:"

    styleN = ParagraphStyle('Normal', spaceAfter=5, leading=10)
    styleH = ParagraphStyle('Heading1', spaceAfter=10, fontSize=16,
                            textColor=colors.darkblue, fontName='Helvetica-Bold')

    # Añadir logotipo y títulos
    logo = Image(logo_path, width=2*inch, height=1*inch)
    logo.hAlign = 'LEFT'
    story.append(logo)
    story.append(Spacer(1, 20))
    story.append(Paragraph(title, styleH))
    story.append(Spacer(1, 12))
    story.append(Paragraph(subtitle, styleN))
    story.append(Spacer(1, 12))

    for usuario in usuarios:
        story.append(Paragraph(f"ID: {usuario.id}", styleN))
        story.append(
            Paragraph(f"Nombre: {usuario.first_name} {usuario.last_name}", styleN))
        story.append(Paragraph(f"Email: {usuario.email}", styleN))
        story.append(
            Paragraph(f"Fecha de Alta: {usuario.date_joined}", styleN))
        story.append(Spacer(1, 12))
        story.append(HRFlowable(width="100%", thickness=1,
                     color=colors.grey, spaceBefore=1, spaceAfter=24))

    doc.build(story)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


def exportar_tareas_excel(request):
    fecha_inicio_desde_tarea = request.GET.get('fecha_inicio_desde_tarea')
    fecha_inicio_hasta_tarea = request.GET.get('fecha_inicio_hasta_tarea')
    tarea_id = request.GET.get('tarea_id')
    titulo_palabras = request.GET.get('titulo_palabras')

    # Filtrar tareas según los criterios de búsqueda
    tareas = TareaPorDesarrollar.objects.all()
    if fecha_inicio_desde_tarea and fecha_inicio_hasta_tarea:
        fecha_inicio_desde_tarea = timezone.make_aware(
            datetime.strptime(fecha_inicio_desde_tarea, '%Y-%m-%d'))
        fecha_inicio_hasta_tarea = timezone.make_aware(
            datetime.strptime(fecha_inicio_hasta_tarea, '%Y-%m-%d'))
        tareas = tareas.filter(fecha_creacion__gte=fecha_inicio_desde_tarea,
                               fecha_creacion__lte=fecha_inicio_hasta_tarea)
    if tarea_id:
        tareas = tareas.filter(id=tarea_id)
    if titulo_palabras:
        tareas = tareas.filter(titulo__icontains=titulo_palabras)

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tareas"

    # Añadir encabezados
    ws.append(["ID", "Título", "Descripción", "Asignado a",
              "Fecha de Creación", "Fecha de Vencimiento", "Estado"])

    # Añadir datos de las tareas
    for tarea in tareas:
        # Convertir datetime a naive (sin zona horaria)
        fecha_creacion = datetime.combine(
            tarea.fecha_creacion, datetime.min.time()) if tarea.fecha_creacion else ''
        fecha_vencimiento = datetime.combine(
            tarea.fecha_vencimiento, datetime.min.time()) if tarea.fecha_vencimiento else ''
        ws.append([tarea.id, tarea.titulo, tarea.descripcion, f"{tarea.usuario.first_name} {tarea.usuario.last_name}",
                   fecha_creacion, fecha_vencimiento, tarea.estado])

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tareas.xlsx'
    wb.save(response)
    return response


def exportar_proyectos_excel(request):
    fecha_inicio_desde = request.GET.get('fecha_inicio_desde')
    fecha_inicio_hasta = request.GET.get('fecha_inicio_hasta')
    proyecto_id = request.GET.get('proyecto_id')
    titulo_palabras = request.GET.get('titulo_palabras')

    # Filtrar proyectos según los criterios de búsqueda
    proyectos = Proyecto.objects.all()
    if fecha_inicio_desde and fecha_inicio_hasta:
        fecha_inicio_desde = timezone.make_aware(
            datetime.strptime(fecha_inicio_desde, '%Y-%m-%d'))
        fecha_inicio_hasta = timezone.make_aware(
            datetime.strptime(fecha_inicio_hasta, '%Y-%m-%d'))
        proyectos = proyectos.filter(
            fecha_inicio__gte=fecha_inicio_desde, fecha_inicio__lte=fecha_inicio_hasta)
    if proyecto_id:
        proyectos = proyectos.filter(id=proyecto_id)
    if titulo_palabras:
        proyectos = proyectos.filter(nombre__icontains=titulo_palabras)

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyectos"

    # Añadir encabezados
    ws.append(["ID", "Nombre", "Descripción",
              "Fecha de Inicio", "Fecha de Fin", "Estado"])

    # Añadir datos de los proyectos
    for proyecto in proyectos:
        # Convertir datetime a naive (sin zona horaria)
        fecha_inicio = datetime.combine(
            proyecto.fecha_inicio, datetime.min.time()) if proyecto.fecha_inicio else ''
        fecha_fin = datetime.combine(
            proyecto.fecha_fin, datetime.min.time()) if proyecto.fecha_fin else ''
        ws.append([proyecto.id, proyecto.nombre, proyecto.descripcion,
                  fecha_inicio, fecha_fin, proyecto.estado])

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proyectos.xlsx'
    wb.save(response)
    return response


def exportar_usuarios_excel(request):
    nombre_o_apellido = request.GET.get('nombre_o_apellido')
    fecha_alta_desde = request.GET.get('fecha_alta_desde')
    fecha_alta_hasta = request.GET.get('fecha_alta_hasta')

    # Filtrar usuarios según los criterios de búsqueda
    usuarios = Usuario.objects.all()
    if nombre_o_apellido:
        usuarios = usuarios.filter(first_name__icontains=nombre_o_apellido) | usuarios.filter(
            last_name__icontains=nombre_o_apellido)
    if fecha_alta_desde and fecha_alta_hasta:
        fecha_alta_desde = timezone.make_aware(
            datetime.strptime(fecha_alta_desde, '%Y-%m-%d'))
        fecha_alta_hasta = timezone.make_aware(
            datetime.strptime(fecha_alta_hasta, '%Y-%m-%d'))
        usuarios = usuarios.filter(
            date_joined__gte=fecha_alta_desde, date_joined__lte=fecha_alta_hasta)

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Añadir encabezados
    ws.append(["ID", "Username", "Nombre", "Apellido",
              "Email", "Fecha de Registro"])

    # Añadir datos de los usuarios filtrados
    for usuario in usuarios:
        # Convertir datetime a naive (sin zona horaria)
        fecha_registro = usuario.date_joined.replace(
            tzinfo=None) if usuario.date_joined else ''
        ws.append([usuario.id, usuario.username, usuario.first_name,
                  usuario.last_name, usuario.email, fecha_registro])

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    wb.save(response)
    return response


def exportar_todos_usuarios_excel(request):
    usuarios = Usuario.objects.all()

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Todos los Usuarios"

    # Añadir encabezados
    ws.append(["ID Usuario", "Username", "Nombre", "Apellido", "Email", "Fecha de Registro",
              "ID Tarea", "Título Tarea", "Estado Tarea", "Fecha de Creación", "Fecha de Vencimiento"])

    # Añadir datos de los usuarios y tareas asignadas
    for usuario in usuarios:
        fecha_registro = usuario.date_joined.replace(tzinfo=None) if usuario.date_joined and hasattr(
            usuario.date_joined, 'tzinfo') else usuario.date_joined
        tareas_asignadas = TareaPorDesarrollar.objects.filter(usuario=usuario)

        if tareas_asignadas:
            for tarea in tareas_asignadas:
                fecha_creacion = tarea.fecha_creacion.replace(tzinfo=None) if tarea.fecha_creacion and hasattr(
                    tarea.fecha_creacion, 'tzinfo') else tarea.fecha_creacion
                fecha_vencimiento = tarea.fecha_vencimiento.replace(tzinfo=None) if tarea.fecha_vencimiento and hasattr(
                    tarea.fecha_vencimiento, 'tzinfo') else tarea.fecha_vencimiento
                ws.append([usuario.id, usuario.username, usuario.first_name, usuario.last_name, usuario.email,
                          fecha_registro, tarea.id, tarea.titulo, tarea.estado, fecha_creacion, fecha_vencimiento])
        else:
            ws.append([usuario.id, usuario.username, usuario.first_name,
                      usuario.last_name, usuario.email, fecha_registro, '', '', '', '', ''])

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=todos_usuarios.xlsx'
    wb.save(response)
    return response


def exportar_todos_proyectos_excel(request):
    proyectos = Proyecto.objects.all()

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Todos los Proyectos"

    # Añadir encabezados
    ws.append(["ID", "Nombre", "Descripción",
              "Fecha de Inicio", "Fecha de Fin", "Estado"])

    # Añadir datos de los proyectos
    for proyecto in proyectos:
        # Convertir datetime a naive (sin zona horaria)
        fecha_inicio = datetime.combine(
            proyecto.fecha_inicio, datetime.min.time()) if proyecto.fecha_inicio else ''
        fecha_fin = datetime.combine(
            proyecto.fecha_fin, datetime.min.time()) if proyecto.fecha_fin else ''
        ws.append([proyecto.id, proyecto.nombre, proyecto.descripcion,
                  fecha_inicio, fecha_fin, proyecto.estado])

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=todos_proyectos.xlsx'
    wb.save(response)
    return response


def exportar_todas_tareas_excel(request):
    tareas = TareaPorDesarrollar.objects.all()

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Todas las Tareas"

    # Añadir encabezados
    ws.append(["ID", "Título", "Descripción", "Estado", "Fecha de Creación",
              "Fecha de Vencimiento", "Proyecto", "Asignado a"])

    # Añadir datos de las tareas
    for tarea in tareas:
        # Convertir datetime a naive (sin zona horaria)
        fecha_creacion = datetime.combine(
            tarea.fecha_creacion, datetime.min.time()) if tarea.fecha_creacion else ''
        fecha_vencimiento = datetime.combine(
            tarea.fecha_vencimiento, datetime.min.time()) if tarea.fecha_vencimiento else ''
        ws.append([tarea.id, tarea.titulo, tarea.descripcion, tarea.estado, fecha_creacion, fecha_vencimiento,
                  tarea.proyecto.nombre, f"{tarea.usuario.first_name} {tarea.usuario.last_name}"])

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=todas_tareas.xlsx'
    wb.save(response)
    return response


def exportar_todas_evaluaciones_excel(request):
    evaluaciones = Evaluacion.objects.all()

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Todas las Evaluaciones"

    # Añadir encabezados
    ws.append(["ID", "Título", "Calificación", "Tarea", "Asignado a",
              "Comentarios", "Fecha de Evaluación", "Proyecto"])

    # Añadir datos de las evaluaciones
    for evaluacion in evaluaciones:
        # Convertir datetime a naive (sin zona horaria)
        fecha_evaluacion = datetime.combine(
            evaluacion.fecha_evaluacion, datetime.min.time()) if evaluacion.fecha_evaluacion else ''
        tarea_titulo = evaluacion.tarea.titulo if evaluacion.tarea else 'Sin Tarea Asignada'
        ws.append([evaluacion.id, evaluacion.titulo, evaluacion.calificacion, tarea_titulo, f"{evaluacion.usuario.first_name} {
                  evaluacion.usuario.last_name}", evaluacion.comentarios, fecha_evaluacion, evaluacion.proyecto.nombre])

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=todas_evaluaciones.xlsx'
    wb.save(response)
    return response
