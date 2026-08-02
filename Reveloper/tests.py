from reportlab.lib import colors
from django.test import TestCase, Client
from django.urls import reverse

from .models import Usuario


class ExportEndpointSecurityTests(TestCase):
    """Pruebas mínimas de autenticación y autorización en exportaciones."""

    GLOBAL_ADMIN_URL_NAMES = [
        'generate_pdf',
        'generate_user_pdf',
        'generar_informe_grafico_pdf_admin',
        'exportar_tareas_excel',
        'exportar_proyectos_excel',
        'exportar_usuarios_excel',
        'exportar_todos_usuarios_excel',
        'exportar_todos_proyectos_excel',
        'exportar_todas_tareas_excel',
        'exportar_todas_evaluaciones_excel',
        'generar_informe_pdf_busqueda',
        'generar_informe_pdf_tareas',
        'generar_informe_pdf_usuarios',
    ]

    DEVELOPER_OWN_URL_NAMES = [
        'generate_task_pdf',
        'generate_evaluation_pdf',
        'generar_informe_grafico_pdf',
    ]

    @classmethod
    def setUpTestData(cls):
        cls.admin = Usuario.objects.create_superuser(
            username='admin_test',
            email='admin@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Test',
        )
        cls.developer = Usuario.objects.create_user(
            username='dev_test',
            email='dev@test.local',
            password='test-pass-123',
            nombre='Dev',
            apellido='Test',
        )

    def setUp(self):
        self.client = Client()

    def _login(self, user):
        self.assertTrue(
            self.client.login(username=user.username, password='test-pass-123')
        )

    def test_anonymous_redirects_to_login_on_global_exports(self):
        for url_name in self.GLOBAL_ADMIN_URL_NAMES:
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 302)
                self.assertIn('/accounts/login/', response.url)

    def test_anonymous_redirects_to_login_on_developer_exports(self):
        for url_name in self.DEVELOPER_OWN_URL_NAMES:
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 302)
                self.assertIn('/accounts/login/', response.url)

    def test_developer_cannot_access_global_exports(self):
        self._login(self.developer)
        for url_name in self.GLOBAL_ADMIN_URL_NAMES:
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 302)
                self.assertIn('/accounts/login/', response.url)

    def test_developer_can_access_own_scoped_exports(self):
        self._login(self.developer)
        for url_name in self.DEVELOPER_OWN_URL_NAMES:
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 200)

    def test_admin_can_access_global_exports(self):
        self._login(self.admin)
        for url_name in self.GLOBAL_ADMIN_URL_NAMES:
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 200)

    def test_print_template_dirs_route_removed(self):
        response = self.client.get('/reveloper/print-template-dirs/')
        self.assertEqual(response.status_code, 404)


class SettingsSecurityTests(TestCase):
    """Pruebas para verificar la externalización de la configuración sensible."""

    def test_debug_conversion_from_string(self):
        test_cases = [
            ('True', True),
            ('true', True),
            ('1', True),
            ('False', False),
            ('false', False),
            ('0', False),
            ('', False),
        ]
        for val, expected in test_cases:
            with self.subTest(val=val):
                res = val.lower() in ('true', '1', 't', 'yes')
                self.assertEqual(res, expected)

    def test_allowed_hosts_parsing(self):
        raw_hosts = 'localhost, 127.0.0.1, example.com '
        parsed = [host.strip() for host in raw_hosts.split(',') if host.strip()]
        self.assertEqual(parsed, ['localhost', '127.0.0.1', 'example.com'])

    def test_test_settings_uses_sqlite_in_memory(self):
        from django.conf import settings
        from django.db import connection

        self.assertEqual(settings.DATABASES['default']['ENGINE'], 'django.db.backends.sqlite3')
        self.assertTrue(
            ':memory:' in settings.DATABASES['default']['NAME'] or 'memory' in settings.DATABASES['default']['NAME']
        )
        self.assertEqual(connection.vendor, 'sqlite')

    def test_env_example_does_not_contain_real_secrets(self):
        from django.conf import settings

        env_example_path = settings.BASE_DIR / '.env.example'
        self.assertTrue(env_example_path.exists(), '.env.example debe existir')
        content = env_example_path.read_text(encoding='utf-8')

        self.assertIn('DJANGO_SECRET_KEY=', content)
        self.assertNotIn('DJANGO_SECRET_KEY=django-insecure', content)
        self.assertNotIn(settings.SECRET_KEY, content)
        self.assertIn('DB_PASSWORD=', content)
    def test_environment_helpers_parse_values(self):
        from unittest.mock import patch

        from ProyectEspecial import settings as project_settings

        with patch.dict(
            'os.environ',
            {
                'TEST_BOOLEAN_VALUE': 'yes',
                'TEST_LIST_VALUE': (
                    'localhost, 127.0.0.1, example.com '
                ),
            },
        ):
            self.assertTrue(
                project_settings._env_bool(
                    'TEST_BOOLEAN_VALUE'
                )
            )
            self.assertEqual(
                project_settings._env_list(
                    'TEST_LIST_VALUE'
                ),
                [
                    'localhost',
                    '127.0.0.1',
                    'example.com',
                ],
            )

    def test_test_environment_does_not_force_https(self):
        from django.conf import settings

        self.assertEqual(settings.ENVIRONMENT, 'testing')
        self.assertFalse(settings.IS_PRODUCTION)
        self.assertFalse(settings.DEBUG)
        self.assertFalse(settings.SECURE_SSL_REDIRECT)
        self.assertFalse(settings.SESSION_COOKIE_SECURE)
        self.assertFalse(settings.CSRF_COOKIE_SECURE)
        self.assertEqual(settings.SECURE_HSTS_SECONDS, 0)

    def test_common_security_headers_are_enabled(self):
        from django.conf import settings

        self.assertTrue(settings.SECURE_CONTENT_TYPE_NOSNIFF)
        self.assertTrue(settings.SESSION_COOKIE_HTTPONLY)
        self.assertEqual(
            settings.SECURE_REFERRER_POLICY,
            'same-origin',
        )
        self.assertEqual(settings.X_FRAME_OPTIONS, 'DENY')

    def test_test_environment_has_valid_allowed_hosts(self):
        from django.conf import settings

        self.assertIn('testserver', settings.ALLOWED_HOSTS)
        self.assertIn('localhost', settings.ALLOWED_HOSTS)

    def test_env_example_documents_security_variables(self):
        from django.conf import settings

        env_example_path = settings.BASE_DIR / '.env.example'
        content = env_example_path.read_text(encoding='utf-8')

        expected_variables = (
            'DJANGO_ENV=',
            'DJANGO_CSRF_TRUSTED_ORIGINS=',
            'DJANGO_SECURE_HSTS_SECONDS=',
            'DJANGO_SECURE_HSTS_INCLUDE_SUBDOMAINS=',
            'DJANGO_SECURE_HSTS_PRELOAD=',
            'DJANGO_USE_X_FORWARDED_PROTO=',
        )

        for variable in expected_variables:
            with self.subTest(variable=variable):
                self.assertIn(variable, content)

    def test_static_files_configuration(self):
        from django.conf import settings

        self.assertEqual(settings.STATIC_URL, '/static/')
        self.assertEqual(
            settings.STATIC_ROOT,
            settings.BASE_DIR / 'staticfiles',
        )
        self.assertEqual(settings.STATICFILES_DIRS, [])

    def test_collected_static_directory_is_ignored(self):
        from django.conf import settings

        gitignore_path = settings.BASE_DIR / '.gitignore'
        content = gitignore_path.read_text(encoding='utf-8')

        ignored_entries = [
            line.strip()
            for line in content.splitlines()
            if line.strip() and not line.lstrip().startswith('#')
        ]

        self.assertIn('staticfiles/', ignored_entries)


class MediaUploadSecurityTests(TestCase):
    """Pruebas que documentan la ausencia de cargas de usuarios."""

    def test_models_do_not_define_file_fields(self):
        from django.db import models

        from .models import (
            Evaluacion,
            EvaluacionConfig,
            Proyecto,
            TareaPorDesarrollar,
            TareasCompletadas,
            Usuario,
        )

        model_classes = (
            Usuario,
            Proyecto,
            TareaPorDesarrollar,
            TareasCompletadas,
            Evaluacion,
            EvaluacionConfig,
        )

        file_field_types = (
            models.FileField,
            models.FilePathField,
        )

        for model_class in model_classes:
            with self.subTest(model=model_class.__name__):
                file_fields = [
                    field
                    for field in model_class._meta.get_fields()
                    if isinstance(field, file_field_types)
                ]

                self.assertEqual(file_fields, [])

    def test_forms_do_not_define_file_upload_fields(self):
        from django import forms

        from .forms import (
            CustomUserChangeForm,
            CustomUserCreationForm,
            EvaluacionForm,
            TareaPorDesarrollarForm,
        )

        form_classes = (
            TareaPorDesarrollarForm,
            CustomUserCreationForm,
            CustomUserChangeForm,
            EvaluacionForm,
        )

        for form_class in form_classes:
            with self.subTest(form=form_class.__name__):
                form = form_class()

                file_fields = [
                    field
                    for field in form.fields.values()
                    if isinstance(field, forms.FileField)
                ]

                self.assertEqual(file_fields, [])

    def test_views_do_not_process_uploaded_files(self):
        from pathlib import Path

        views_path = Path(__file__).resolve().parent / 'views.py'
        source = views_path.read_text(encoding='utf-8')

        forbidden_patterns = (
            'request.FILES',
            'FileSystemStorage',
            'default_storage',
            'UploadedFile',
        )

        for pattern in forbidden_patterns:
            with self.subTest(pattern=pattern):
                self.assertNotIn(pattern, source)

    def test_templates_do_not_contain_file_upload_inputs(self):
        from pathlib import Path

        templates_path = (
            Path(__file__).resolve().parent / 'templates'
        )

        forbidden_patterns = (
            'multipart/form-data',
            'type="file"',
            "type='file'",
        )

        for template_path in templates_path.rglob('*.html'):
            source = template_path.read_text(encoding='utf-8')

            for pattern in forbidden_patterns:
                with self.subTest(
                    template=template_path.name,
                    pattern=pattern,
                ):
                    self.assertNotIn(pattern, source)


class ExportResourcePathTests(TestCase):
    """Pruebas para recursos usados en informes exportados."""

    def test_report_logo_is_resolved_as_existing_absolute_path(self):
        from pathlib import Path

        from .views import _obtener_ruta_logo_informes

        logo_path = Path(_obtener_ruta_logo_informes())

        self.assertTrue(logo_path.is_absolute())
        self.assertTrue(logo_path.exists())
        self.assertEqual(
            logo_path.name,
            'logo-reveloper.png',
        )

    def test_report_logo_resolution_does_not_depend_on_cwd(self):
        import os
        import tempfile
        from pathlib import Path

        from .views import _obtener_ruta_logo_informes

        original_cwd = Path.cwd()

        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                os.chdir(temp_dir)

                logo_path = Path(
                    _obtener_ruta_logo_informes()
                )

                self.assertTrue(logo_path.is_absolute())
                self.assertTrue(logo_path.exists())
            finally:
                os.chdir(original_cwd)

    def test_missing_report_logo_raises_clear_error(self):
        from unittest.mock import patch

        from .views import _obtener_ruta_logo_informes

        with patch(
            'Reveloper.views.finders.find',
            return_value=None,
        ):
            with self.assertRaisesRegex(
                FileNotFoundError,
                'No se encontró el logotipo',
            ):
                _obtener_ruta_logo_informes()

    def test_views_do_not_use_relative_static_logo_path(self):
        from pathlib import Path

        views_path = Path(__file__).resolve().parent / 'views.py'
        source = views_path.read_text(encoding='utf-8')

        self.assertNotIn(
            'Reveloper/static/img/logos/logo-reveloper.png',
            source,
        )
        self.assertEqual(
            source.count(
                'logo_path = _obtener_ruta_logo_informes()'
            ),
            7,
        )


class UserCreationFormTests(TestCase):
    """Pruebas del formulario personalizado de creación de usuarios."""

    def test_creation_form_includes_custom_name_fields(self):
        from .forms import CustomUserCreationForm

        form = CustomUserCreationForm()

        self.assertIn('nombre', form.fields)
        self.assertIn('apellido', form.fields)

    def test_custom_name_fields_are_required(self):
        from .forms import CustomUserCreationForm

        form = CustomUserCreationForm()

        self.assertTrue(form.fields['nombre'].required)
        self.assertTrue(form.fields['apellido'].required)

    def test_form_without_custom_name_fields_is_invalid(self):
        from .forms import CustomUserCreationForm

        form = CustomUserCreationForm(
            data={
                'username': 'usuario_sin_nombre',
                'email': 'sin.nombre@test.local',
                'password1': 'test-pass-123',
                'password2': 'test-pass-123',
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn('nombre', form.errors)
        self.assertIn('apellido', form.errors)

    def test_valid_form_saves_custom_name_fields(self):
        from .forms import CustomUserCreationForm

        form = CustomUserCreationForm(
            data={
                'username': 'usuario_formulario',
                'email': 'usuario.formulario@test.local',
                'nombre': 'Roberto',
                'apellido': 'González',
                'password1': 'test-pass-123',
                'password2': 'test-pass-123',
            }
        )

        self.assertTrue(form.is_valid(), form.errors)

        usuario = form.save()

        self.assertEqual(usuario.nombre, 'Roberto')
        self.assertEqual(usuario.apellido, 'González')
        self.assertEqual(
            usuario.email,
            'usuario.formulario@test.local',
        )
        self.assertTrue(
            usuario.check_password('test-pass-123')
        )

class CustomUserAdminTests(TestCase):
    """Pruebas de integración del usuario personalizado con Django Admin."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = Usuario.objects.create_superuser(
            username='admin_user_admin',
            email='admin.user@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Usuarios',
        )

    def setUp(self):
        self.client = Client()
        self.assertTrue(
            self.client.login(
                username=self.admin.username,
                password='test-pass-123',
            )
        )

    def test_custom_user_admin_uses_custom_forms(self):
        from django.contrib import admin

        from .admin import CustomUserAdmin
        from .forms import (
            CustomUserChangeForm,
            CustomUserCreationForm,
        )

        registered_admin = admin.site._registry[Usuario]

        self.assertIsInstance(
            registered_admin,
            CustomUserAdmin,
        )
        self.assertIs(
            registered_admin.add_form,
            CustomUserCreationForm,
        )
        self.assertIs(
            registered_admin.form,
            CustomUserChangeForm,
        )

    def test_admin_add_fieldsets_include_custom_name_fields(self):
        from django.contrib import admin

        registered_admin = admin.site._registry[Usuario]
        add_fields = registered_admin.add_fieldsets[0][1]['fields']

        self.assertIn('nombre', add_fields)
        self.assertIn('apellido', add_fields)

    def test_admin_change_fieldsets_include_custom_name_fields(self):
        from django.contrib import admin

        registered_admin = admin.site._registry[Usuario]

        all_fields = {
            field
            for _, options in registered_admin.fieldsets
            for field in options['fields']
        }

        self.assertIn('nombre', all_fields)
        self.assertIn('apellido', all_fields)

    def test_admin_add_page_displays_custom_name_fields(self):
        app_label = Usuario._meta.app_label
        model_name = Usuario._meta.model_name
        url = reverse(
            f'admin:{app_label}_{model_name}_add'
        )

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="nombre"')
        self.assertContains(response, 'name="apellido"')

    def test_admin_change_page_displays_custom_name_fields(self):
        app_label = Usuario._meta.app_label
        model_name = Usuario._meta.model_name
        url = reverse(
            f'admin:{app_label}_{model_name}_change',
            args=[self.admin.pk],
        )

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="nombre"')
        self.assertContains(response, 'name="apellido"')

class TaskLifecycleTests(TestCase):
    """Pruebas automatizadas para el ciclo de vida de tareas (Hito 4)."""

    @classmethod
    def setUpTestData(cls):
        from datetime import date
        from .models import Proyecto, TareaPorDesarrollar, TareasCompletadas

        cls.admin = Usuario.objects.create_superuser(
            username='admin_lifecycle',
            email='admin_lifecycle@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Lifecycle',
        )
        cls.developer = Usuario.objects.create_user(
            username='dev_lifecycle',
            email='dev_lifecycle@test.local',
            password='test-pass-123',
            nombre='Dev',
            apellido='Lifecycle',
        )
        cls.staff_user = Usuario.objects.create_user(
            username='staff_lifecycle',
            email='staff_lifecycle@test.local',
            password='test-pass-123',
            nombre='Staff',
            apellido='Lifecycle',
            is_staff=True,
        )
        cls.inactive_user = Usuario.objects.create_user(
            username='inactive_lifecycle',
            email='inactive_lifecycle@test.local',
            password='test-pass-123',
            nombre='Inactive',
            apellido='Lifecycle',
            is_active=False,
        )
        cls.proyecto = Proyecto.objects.create(
            id='PROY-001',
            nombre='Proyecto Prueba',
            descripcion='Descripción del proyecto',
            fecha_inicio=date.today(),
            fecha_fin=date.today(),
            estado='activo',
        )
        cls.tarea_text_id = 'TASK-TEXT-1001'
        cls.tarea = TareaPorDesarrollar.objects.create(
            id=cls.tarea_text_id,
            usuario=cls.developer,
            proyecto=cls.proyecto,
            titulo='Tarea Texto Inicial',
            descripcion='Descripción inicial',
            fecha_vencimiento=date.today(),
            estado='pendiente',
        )


    def setUp(self):
        self.client = Client()

    def _login(self, user):
        self.assertTrue(
            self.client.login(username=user.username, password='test-pass-123')
        )

    def test_editar_tarea_accepts_string_id(self):
        self._login(self.admin)
        url = reverse('editar_tarea', kwargs={'tarea_id': self.tarea_text_id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

    def test_anonymous_user_cannot_edit_task(self):
        url = reverse('editar_tarea', kwargs={'tarea_id': self.tarea_text_id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/accounts/login/', response.url)

    def test_developer_cannot_edit_task(self):
        self._login(self.developer)
        url = reverse('editar_tarea', kwargs={'tarea_id': self.tarea_text_id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/accounts/login/', response.url)

    def test_admin_can_open_edit_task_get(self):
        self._login(self.admin)
        url = reverse('editar_tarea', kwargs={'tarea_id': self.tarea_text_id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'editar_tarea.html')

    def test_valid_post_modifies_task(self):
        from datetime import date
        self._login(self.admin)
        url = reverse('editar_tarea', kwargs={'tarea_id': self.tarea_text_id})
        post_data = {
            'titulo': 'Tarea Texto Modificada',
            'descripcion': 'Descripción modificada',
            'fecha_vencimiento': str(date.today()),
            'estado': 'en progreso',
            'proyecto': self.proyecto.id,
            'usuario': self.developer.id,
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.titulo, 'Tarea Texto Modificada')
        self.assertEqual(self.tarea.estado, 'en progreso')

    def test_invalid_post_shows_errors_and_does_not_save(self):
        from datetime import date
        self._login(self.admin)
        url = reverse('editar_tarea', kwargs={'tarea_id': self.tarea_text_id})
        post_data = {
            'titulo': '',
            'descripcion': 'Sin título',
            'fecha_vencimiento': str(date.today()),
            'estado': 'pendiente',
            'proyecto': self.proyecto.id,
            'usuario': self.developer.id,
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        self.assertIn('titulo', response.context['form'].errors)
        self.tarea.refresh_from_db()
        self.assertNotEqual(self.tarea.descripcion, 'Sin título')

    def test_crear_tarea_assigns_to_selected_developer(self):
        from datetime import date
        from .models import TareaPorDesarrollar
        self._login(self.admin)
        url = reverse('crear_tarea')
        post_data = {
            'id': 'TASK-NEW-200',
            'titulo': 'Tarea para desarrollador',
            'descripcion': 'Detalle tarea',
            'fecha_vencimiento': str(date.today()),
            'estado': 'pendiente',
            'proyecto': self.proyecto.id,
            'usuario': self.developer.id,
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        nueva_tarea = TareaPorDesarrollar.objects.filter(titulo='Tarea para desarrollador').first()
        self.assertIsNotNone(nueva_tarea)
        self.assertEqual(nueva_tarea.usuario, self.developer)

    def test_completing_task_creates_exactly_one_tareas_completadas(self):
        from datetime import date
        from .models import TareaPorDesarrollar, TareasCompletadas
        tarea_to_complete = TareaPorDesarrollar.objects.create(
            id='TASK-COMPLETABLE-1',
            usuario=self.developer,
            proyecto=self.proyecto,
            titulo='Tarea a completar',
            fecha_vencimiento=date.today(),
            estado='pendiente',
        )
        self.assertEqual(TareasCompletadas.objects.filter(tarea_original_id=tarea_to_complete.id).count(), 0)

        tarea_to_complete.estado = 'completada'
        tarea_to_complete.save()

        count = TareasCompletadas.objects.filter(tarea_original_id=tarea_to_complete.id).count()
        self.assertEqual(count, 1)

    def test_completing_task_increments_tareas_completadas_counter_once(self):
        from datetime import date
        from .models import TareaPorDesarrollar
        initial_counter = Usuario.objects.get(pk=self.developer.pk).tareas_completadas
        tarea_to_complete = TareaPorDesarrollar.objects.create(
            id='TASK-COMPLETABLE-2',
            usuario=self.developer,
            proyecto=self.proyecto,
            titulo='Tarea contador',
            fecha_vencimiento=date.today(),
            estado='pendiente',
        )

        tarea_to_complete.estado = 'completada'
        tarea_to_complete.save()

        updated_counter = Usuario.objects.get(pk=self.developer.pk).tareas_completadas
        self.assertEqual(updated_counter, initial_counter + 1)

    def test_resaving_completed_task_does_not_duplicate_history_or_counter(self):
        from datetime import date
        from .models import TareaPorDesarrollar, TareasCompletadas
        tarea_completed = TareaPorDesarrollar.objects.create(
            id='TASK-COMPLETABLE-3',
            usuario=self.developer,
            proyecto=self.proyecto,
            titulo='Tarea a resguardar',
            fecha_vencimiento=date.today(),
            estado='pendiente',
        )
        tarea_completed.estado = 'completada'
        tarea_completed.save()

        history_count_before = TareasCompletadas.objects.filter(tarea_original_id=tarea_completed.id).count()
        counter_before = Usuario.objects.get(pk=self.developer.pk).tareas_completadas

        tarea_completed.descripcion = 'Descripción actualizada post completado'
        tarea_completed.save()

        history_count_after = TareasCompletadas.objects.filter(tarea_original_id=tarea_completed.id).count()
        counter_after = Usuario.objects.get(pk=self.developer.pk).tareas_completadas

        self.assertEqual(history_count_after, history_count_before)
        self.assertEqual(counter_after, counter_before)

    def test_pending_task_does_not_create_completed_history(self):
        from datetime import date
        from .models import TareaPorDesarrollar, TareasCompletadas

        tarea = TareaPorDesarrollar.objects.create(
            id='TASK-PENDING-NO-HISTORY',
            usuario=self.developer,
            proyecto=self.proyecto,
            titulo='Tarea pendiente sin historial',
            fecha_vencimiento=date.today(),
            estado='pendiente',
        )

        self.assertFalse(
            TareasCompletadas.objects.filter(
                tarea_original_id=tarea.id
            ).exists()
        )

    def test_completed_history_preserves_basic_task_data(self):
        from datetime import date
        from django.utils import timezone
        from .models import TareaPorDesarrollar, TareasCompletadas

        fecha_vencimiento = date.today()
        tarea = TareaPorDesarrollar.objects.create(
            id='TASK-HISTORY-DATA',
            usuario=self.developer,
            proyecto=self.proyecto,
            titulo='Tarea con datos históricos',
            fecha_vencimiento=fecha_vencimiento,
            estado='pendiente',
        )

        tarea.estado = 'completada'
        tarea.save()

        historial = TareasCompletadas.objects.get(
            tarea_original_id=tarea.id
        )

        self.assertEqual(historial.titulo, tarea.titulo)
        self.assertEqual(historial.estado, 'completada')
        self.assertEqual(historial.usuario, self.developer)
        self.assertEqual(
            timezone.localdate(historial.fecha_entrega),
            fecha_vencimiento,
        )

    def test_changing_completed_task_to_another_state_does_not_increment_again(self):
        from datetime import date
        from .models import TareaPorDesarrollar, TareasCompletadas

        tarea = TareaPorDesarrollar.objects.create(
            id='TASK-COMPLETED-REVERTED',
            usuario=self.developer,
            proyecto=self.proyecto,
            titulo='Tarea completada y revertida',
            fecha_vencimiento=date.today(),
            estado='pendiente',
        )

        initial_counter = Usuario.objects.get(
            pk=self.developer.pk
        ).tareas_completadas

        tarea.estado = 'completada'
        tarea.save()

        counter_after_completion = Usuario.objects.get(
            pk=self.developer.pk
        ).tareas_completadas

        tarea.estado = 'en revision'
        tarea.save()

        final_counter = Usuario.objects.get(
            pk=self.developer.pk
        ).tareas_completadas

        self.assertEqual(
            counter_after_completion,
            initial_counter + 1,
        )
        self.assertEqual(
            final_counter,
            counter_after_completion,
        )
        self.assertEqual(
            TareasCompletadas.objects.filter(
                tarea_original_id=tarea.id
            ).count(),
            1,
        )

    def test_superuser_not_in_user_form_queryset(self):
        from .forms import TareaPorDesarrollarForm
        form = TareaPorDesarrollarForm()
        self.assertNotIn(self.admin, form.fields['usuario'].queryset)

    def test_staff_user_not_in_user_form_queryset(self):
        from .forms import TareaPorDesarrollarForm
        form = TareaPorDesarrollarForm()
        self.assertNotIn(self.staff_user, form.fields['usuario'].queryset)

    def test_active_normal_user_in_user_form_queryset(self):
        from .forms import TareaPorDesarrollarForm
        form = TareaPorDesarrollarForm()
        self.assertIn(self.developer, form.fields['usuario'].queryset)

    def test_inactive_user_not_in_user_form_queryset(self):
        from .forms import TareaPorDesarrollarForm
        form = TareaPorDesarrollarForm()
        self.assertNotIn(self.inactive_user, form.fields['usuario'].queryset)

    def test_post_without_user_does_not_create_task(self):
        from datetime import date
        from .models import TareaPorDesarrollar
        self._login(self.admin)
        url = reverse('crear_tarea')
        post_data = {
            'id': 'TASK-NO-USER',
            'titulo': 'Tarea sin usuario',
            'descripcion': 'Detalle',
            'fecha_vencimiento': str(date.today()),
            'estado': 'pendiente',
            'proyecto': self.proyecto.id,
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        self.assertIn('usuario', response.context['form'].errors)
        self.assertFalse(TareaPorDesarrollar.objects.filter(id='TASK-NO-USER').exists())

    def test_post_assigning_staff_user_does_not_create_task(self):
        from datetime import date
        from .models import TareaPorDesarrollar
        self._login(self.admin)
        url = reverse('crear_tarea')
        post_data = {
            'id': 'TASK-STAFF-USER',
            'titulo': 'Tarea para staff',
            'descripcion': 'Detalle',
            'fecha_vencimiento': str(date.today()),
            'estado': 'pendiente',
            'proyecto': self.proyecto.id,
            'usuario': self.staff_user.id,
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        self.assertIn('usuario', response.context['form'].errors)
        self.assertFalse(TareaPorDesarrollar.objects.filter(id='TASK-STAFF-USER').exists())

    def test_get_cannot_change_task_state(self):
        self._login(self.developer)
        url = reverse(
            'marcar_tarea_en_revision',
            kwargs={'tarea_id': self.tarea.id},
        )

        response = self.client.get(url)

        self.assertEqual(response.status_code, 405)
        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.estado, 'pendiente')

    def test_post_changes_pending_task_to_in_progress(self):
        self._login(self.developer)
        url = reverse(
            'marcar_tarea_en_revision',
            kwargs={'tarea_id': self.tarea.id},
        )

        response = self.client.post(url)

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(
            response,
            reverse('tareas_por_desarrollar'),
        )
        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.estado, 'en progreso')

    def test_post_changes_in_progress_task_to_in_review(self):
        self.tarea.estado = 'en progreso'
        self.tarea.save(update_fields=['estado'])

        self._login(self.developer)
        url = reverse(
            'marcar_tarea_en_revision',
            kwargs={'tarea_id': self.tarea.id},
        )

        response = self.client.post(url)

        self.assertEqual(response.status_code, 302)
        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.estado, 'en revision')

    def test_other_developer_cannot_change_task_state(self):
        other_developer = Usuario.objects.create_user(
            username='other_task_developer',
            email='other.task@test.local',
            password='test-pass-123',
            nombre='Other',
            apellido='Developer',
        )
        self._login(other_developer)

        url = reverse(
            'marcar_tarea_en_revision',
            kwargs={'tarea_id': self.tarea.id},
        )
        response = self.client.post(url)

        self.assertEqual(response.status_code, 302)
        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.estado, 'pendiente')

    def test_completed_task_state_is_not_changed(self):
        self.tarea.estado = 'completada'
        self.tarea.save(update_fields=['estado'])

        self._login(self.developer)
        url = reverse(
            'marcar_tarea_en_revision',
            kwargs={'tarea_id': self.tarea.id},
        )

        response = self.client.post(url)

        self.assertEqual(response.status_code, 302)
        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.estado, 'completada')


class EvaluationFlowTests(TestCase):
    """Pruebas para el Hito 5: unificar y corregir el flujo de evaluaciones."""

    @classmethod
    def setUpTestData(cls):
        from datetime import date
        from django.utils import timezone
        from .models import Proyecto, Evaluacion

        cls.admin = Usuario.objects.create_superuser(
            username='admin_eval',
            email='admin_eval@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Eval',
        )
        cls.developer = Usuario.objects.create_user(
            username='dev_eval',
            email='dev_eval@test.local',
            password='test-pass-123',
            nombre='Dev',
            apellido='Eval',
        )
        cls.other_developer = Usuario.objects.create_user(
            username='other_dev_eval',
            email='other_dev_eval@test.local',
            password='test-pass-123',
            nombre='Other',
            apellido='Eval',
        )
        cls.proyecto = Proyecto.objects.create(
            id='PROY-EVAL-001',
            nombre='Proyecto Evaluaciones',
            descripcion='Para pruebas de evaluaciones',
            fecha_inicio=date.today(),
            fecha_fin=date.today(),
            estado='activo',
        )
        cls.evaluacion = Evaluacion.objects.create(
            titulo='Evaluación de prueba',
            comentarios='Comentario de prueba',
            fecha_evaluacion=timezone.now(),
            proyecto=cls.proyecto,
            usuario=cls.developer,
            calificacion=80,
        )
        cls.other_evaluacion = Evaluacion.objects.create(
            titulo='Evaluación de other_developer',
            comentarios='Comentario other',
            fecha_evaluacion=timezone.now(),
            proyecto=cls.proyecto,
            usuario=cls.other_developer,
            calificacion=70,
        )

    def setUp(self):
        self.client = Client()

    def _login(self, user):
        self.assertTrue(
            self.client.login(username=user.username, password='test-pass-123')
        )

    # T1 — Existe exactamente una entrada en urlpatterns con name='evaluaciones'
    def test_single_evaluaciones_route_exists(self):
        from Reveloper.urls import urlpatterns
        matches = [p for p in urlpatterns if getattr(p, 'name', None) == 'evaluaciones']
        self.assertEqual(
            len(matches), 1,
            f"Se esperaba exactamente 1 ruta con name='evaluaciones', se encontraron {len(matches)}",
        )

    # T2 — La URL principal resuelve a la vista `evaluaciones`
    def test_evaluaciones_url_resolves_to_correct_view(self):
        from django.urls import resolve
        match = resolve('/reveloper/evaluaciones/')
        self.assertEqual(match.func.__name__, 'evaluaciones')

    # T3 — Un usuario anónimo es redirigido al login
    def test_anonymous_user_redirected_to_login(self):
        response = self.client.get(reverse('evaluaciones'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/accounts/login/', response.url)

    # T4 — El administrador ve ambas evaluaciones (global)
    def test_admin_sees_all_evaluaciones(self):
        self._login(self.admin)
        response = self.client.get(reverse('evaluaciones'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'evaluaciones.html')
        qs = response.context['evaluaciones']
        self.assertIn(self.evaluacion, qs)
        self.assertIn(self.other_evaluacion, qs)

    # T5 — developer ve su propia evaluación y no la de other_developer
    def test_developer_sees_only_own_evaluaciones(self):
        self._login(self.developer)
        response = self.client.get(reverse('evaluaciones'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'evaluaciones.html')
        qs = response.context['evaluaciones']
        self.assertIn(self.evaluacion, qs)
        self.assertNotIn(self.other_evaluacion, qs)

    # T6 — No existe ruta parametrizada de evaluaciones por ID que exponga datos ajenos
    def test_no_parametrized_evaluaciones_route(self):
        from django.urls import NoReverseMatch
        try:
            reverse('evaluaciones', kwargs={'user_id': self.other_developer.pk})
            self.fail('No debe existir ruta evaluaciones parametrizada por ID de usuario')
        except NoReverseMatch:
            pass  # Correcto: la ruta no existe parametrizada

    # T7 — other_developer ve únicamente su propia evaluación
    def test_other_developer_sees_only_own_evaluaciones(self):
        self._login(self.other_developer)
        response = self.client.get(reverse('evaluaciones'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'evaluaciones.html')
        qs = response.context['evaluaciones']
        self.assertIn(self.other_evaluacion, qs)
        self.assertNotIn(self.evaluacion, qs)

    # T8 — developer no puede ver evaluaciones de other_developer (aislamiento)
    def test_developer_cannot_see_other_developer_evaluaciones(self):
        self._login(self.developer)
        response = self.client.get(reverse('evaluaciones'))
        self.assertEqual(response.status_code, 200)
        qs = response.context['evaluaciones']
        ids_visible = list(qs.values_list('usuario_id', flat=True))
        for uid in ids_visible:
            self.assertEqual(uid, self.developer.pk,
                             'developer no debe ver evaluaciones de otro usuario')

    # T9 — generar_grafico_evaluaciones recibe un desarrollador válido (no un request)
    def test_generar_grafico_evaluaciones_requires_desarrollador_arg(self):
        import inspect
        from .views import generar_grafico_evaluaciones
        sig = inspect.signature(generar_grafico_evaluaciones)
        params = list(sig.parameters.keys())
        self.assertIn('desarrollador', params,
                      'generar_grafico_evaluaciones debe aceptar el parámetro desarrollador')
        self.assertEqual(params[0], 'request')
        self.assertEqual(params[1], 'desarrollador')

    # T10 — No queda ninguna llamada con firma incorrecta (sin desarrollador)
    def test_no_incompatible_calls_to_graph_function(self):
        """Verifica que no existe ninguna llamada a generar_grafico_evaluaciones(request)
        sin el argumento desarrollador en views.py."""
        import re
        import os
        views_path = os.path.join(os.path.dirname(__file__), 'views.py')
        with open(views_path, encoding='utf-8') as f:
            source = f.read()
        bad_calls = re.findall(
            r'generar_grafico_evaluaciones\s*\(\s*request\s*\)',
            source
        )
        self.assertEqual(bad_calls, [],
                         f'Llamadas incompatibles encontradas: {bad_calls}')

    # T11 — El nombre vista_evaluaciones no existe en el sistema de rutas
    def test_vista_evaluaciones_route_does_not_exist(self):
        from django.urls import NoReverseMatch
        with self.assertRaises(NoReverseMatch):
            reverse('vista_evaluaciones')

class TaskReviewEvaluationTests(TestCase):
    """Pruebas del proceso administrativo de revisión de tareas."""

    @classmethod
    def setUpTestData(cls):
        from datetime import date

        from .models import (
            EvaluacionConfig,
            Proyecto,
            TareaPorDesarrollar,
        )

        cls.admin = Usuario.objects.create_superuser(
            username='admin_task_review',
            email='admin.task.review@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Review',
        )
        cls.developer = Usuario.objects.create_user(
            username='developer_task_review',
            email='developer.task.review@test.local',
            password='test-pass-123',
            nombre='Developer',
            apellido='Review',
        )
        cls.proyecto = Proyecto.objects.create(
            id='PROY-REVIEW-001',
            nombre='Proyecto Review',
            descripcion='Proyecto para revisar tareas',
            fecha_inicio=date.today(),
            fecha_fin=date.today(),
            estado='activo',
        )
        cls.config = EvaluacionConfig.objects.create(
            tiempo_entrega=25,
            complejidad_tarea=25,
            cumplimiento_requerimientos=25,
            calidad_codigo=25,
            nota_maxima=100,
        )
        cls.tarea = TareaPorDesarrollar.objects.create(
            id='TASK-REVIEW-001',
            titulo='Tarea en revisión',
            descripcion='Tarea para evaluar',
            fecha_vencimiento=date.today(),
            estado='en revision',
            proyecto=cls.proyecto,
            usuario=cls.developer,
        )

    def setUp(self):
        self.client = Client()

    def _login(self, user):
        self.assertTrue(
            self.client.login(
                username=user.username,
                password='test-pass-123',
            )
        )

    def _valid_post_data(self):
        return {
            'tarea_id': self.tarea.id,
            'tiempo_entrega': '25.0',
            'complejidad_tarea': '25.0',
            'cumplimiento_requerimientos': '25.0',
            'calidad_codigo': '25.0',
        }

    def test_anonymous_user_cannot_review_tasks(self):
        response = self.client.get(reverse('revisar_tareas'))

        self.assertEqual(response.status_code, 302)
        self.assertIn('/accounts/login/', response.url)

    def test_developer_cannot_review_tasks(self):
        self._login(self.developer)

        response = self.client.get(reverse('revisar_tareas'))

        self.assertEqual(response.status_code, 302)

    def test_valid_review_creates_evaluation_and_completes_task(self):
        from .models import Evaluacion

        self._login(self.admin)

        response = self.client.post(
            reverse('revisar_tareas'),
            self._valid_post_data(),
        )

        self.assertEqual(response.status_code, 302)

        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.estado, 'completada')

        evaluacion = Evaluacion.objects.get(tarea=self.tarea)
        self.assertEqual(evaluacion.calificacion, 100)
        self.assertEqual(evaluacion.usuario, self.developer)
        self.assertEqual(evaluacion.proyecto, self.proyecto)

    def test_negative_score_does_not_complete_task(self):
        from .models import Evaluacion

        self._login(self.admin)
        post_data = self._valid_post_data()
        post_data['tiempo_entrega'] = '-1.0'

        response = self.client.post(
            reverse('revisar_tareas'),
            post_data,
        )

        self.assertEqual(response.status_code, 200)

        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.estado, 'en revision')
        self.assertFalse(
            Evaluacion.objects.filter(tarea=self.tarea).exists()
        )

    def test_score_above_configured_limit_is_invalid(self):
        from .models import Evaluacion

        self._login(self.admin)
        post_data = self._valid_post_data()
        post_data['calidad_codigo'] = '26.0'

        response = self.client.post(
            reverse('revisar_tareas'),
            post_data,
        )

        self.assertEqual(response.status_code, 200)

        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.estado, 'en revision')
        self.assertFalse(
            Evaluacion.objects.filter(tarea=self.tarea).exists()
        )

    def test_task_outside_review_state_cannot_be_evaluated(self):
        self.tarea.estado = 'en progreso'
        self.tarea.save(update_fields=['estado'])

        self._login(self.admin)

        response = self.client.post(
            reverse('revisar_tareas'),
            self._valid_post_data(),
        )

        self.assertEqual(response.status_code, 404)

    def test_existing_evaluation_prevents_duplicate(self):
        from django.utils import timezone

        from .models import Evaluacion

        Evaluacion.objects.create(
            titulo='Evaluación existente',
            comentarios='Ya evaluada',
            fecha_evaluacion=timezone.now(),
            proyecto=self.proyecto,
            usuario=self.developer,
            calificacion=100,
            tarea=self.tarea,
        )

        self._login(self.admin)

        response = self.client.post(
            reverse('revisar_tareas'),
            self._valid_post_data(),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'Esta tarea ya tiene una evaluación registrada.',
        )
        self.assertEqual(
            Evaluacion.objects.filter(tarea=self.tarea).count(),
            1,
        )

        self.tarea.refresh_from_db()
        self.assertEqual(self.tarea.estado, 'en revision')

class EvaluationTaskUniquenessTests(TestCase):
    """Pruebas de integridad entre evaluaciones y tareas."""

    @classmethod
    def setUpTestData(cls):
        from datetime import date

        from .models import Proyecto, TareaPorDesarrollar

        cls.developer = Usuario.objects.create_user(
            username='developer_unique_eval',
            email='developer.unique.eval@test.local',
            password='test-pass-123',
            nombre='Developer',
            apellido='Unique',
        )
        cls.proyecto = Proyecto.objects.create(
            id='PROY-UNIQUE-EVAL',
            nombre='Proyecto evaluación única',
            descripcion='Pruebas de integridad',
            fecha_inicio=date.today(),
            fecha_fin=date.today(),
            estado='activo',
        )
        cls.tarea = TareaPorDesarrollar.objects.create(
            id='TASK-UNIQUE-EVAL',
            titulo='Tarea evaluación única',
            descripcion='Solo debe admitir una evaluación',
            fecha_vencimiento=date.today(),
            estado='en revision',
            proyecto=cls.proyecto,
            usuario=cls.developer,
        )

    def _create_evaluation(self, *, titulo, tarea):
        from django.utils import timezone

        from .models import Evaluacion

        return Evaluacion.objects.create(
            titulo=titulo,
            comentarios='Evaluación de prueba',
            fecha_evaluacion=timezone.now(),
            proyecto=self.proyecto,
            usuario=self.developer,
            calificacion=100,
            tarea=tarea,
        )

    def test_evaluation_task_field_is_one_to_one(self):
        from django.db.models import OneToOneField

        from .models import Evaluacion

        field = Evaluacion._meta.get_field('tarea')

        self.assertIsInstance(field, OneToOneField)
        self.assertTrue(field.null)
        self.assertTrue(field.blank)

    def test_task_accepts_one_evaluation(self):
        from .models import Evaluacion

        evaluacion = self._create_evaluation(
            titulo='Primera evaluación',
            tarea=self.tarea,
        )

        self.assertEqual(
            Evaluacion.objects.filter(tarea=self.tarea).count(),
            1,
        )
        self.assertEqual(evaluacion.tarea, self.tarea)

    def test_second_evaluation_for_same_task_raises_integrity_error(self):
        from django.db import IntegrityError, transaction

        self._create_evaluation(
            titulo='Primera evaluación',
            tarea=self.tarea,
        )

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self._create_evaluation(
                    titulo='Segunda evaluación',
                    tarea=self.tarea,
                )

    def test_multiple_evaluations_without_task_are_allowed(self):
        from .models import Evaluacion

        self._create_evaluation(
            titulo='Evaluación histórica uno',
            tarea=None,
        )
        self._create_evaluation(
            titulo='Evaluación histórica dos',
            tarea=None,
        )

        self.assertEqual(
            Evaluacion.objects.filter(tarea__isnull=True).count(),
            2,
        )

    def test_reverse_relation_returns_single_evaluation(self):
        evaluacion = self._create_evaluation(
            titulo='Evaluación inversa',
            tarea=self.tarea,
        )

        self.assertEqual(self.tarea.evaluacion, evaluacion)

class SearchFilterModuleTests(TestCase):
    """Pruebas estructurales del módulo compartido de filtros."""

    def test_filter_helpers_live_in_filters_module(self):
        from .filters import (
            _parsear_fecha_filtro,
            _filtrar_usuarios,
            _filtrar_proyectos,
            _filtrar_tareas,
        )

        helpers = (
            _parsear_fecha_filtro,
            _filtrar_usuarios,
            _filtrar_proyectos,
            _filtrar_tareas,
        )

        for helper in helpers:
            with self.subTest(helper=helper.__name__):
                self.assertEqual(
                    helper.__module__,
                    'Reveloper.filters',
                )

    def test_views_import_filter_helpers(self):
        from pathlib import Path

        views_path = Path(__file__).resolve().parent / 'views.py'
        source = views_path.read_text(encoding='utf-8')

        self.assertIn('from .filters import (', source)
        self.assertNotIn(
            'def _parsear_fecha_filtro',
            source,
        )
        self.assertNotIn(
            'def _filtrar_usuarios',
            source,
        )
        self.assertNotIn(
            'def _filtrar_proyectos',
            source,
        )
        self.assertNotIn(
            'def _filtrar_tareas',
            source,
        )


class ProjectTaskDateFilterTests(TestCase):
    """Pruebas del Hito 9: filtros seguros de proyectos y tareas."""

    @classmethod
    def setUpTestData(cls):
        from datetime import date, datetime
        from django.utils import timezone
        from .models import Proyecto, TareaPorDesarrollar

        cls.admin = Usuario.objects.create_superuser(
            username='admin_date_filters',
            email='admin_date_filters@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Filtros',
        )
        cls.developer = Usuario.objects.create_user(
            username='developer_date_filters',
            email='developer_date_filters@test.local',
            password='test-pass-123',
            nombre='Developer',
            apellido='Filtros',
        )

        cls.proyecto_enero = Proyecto.objects.create(
            id='PROY-DATE-001',
            nombre='Proyecto Enero',
            descripcion='Proyecto de enero',
            fecha_inicio=date(2026, 1, 10),
            fecha_fin=date(2026, 1, 31),
            estado='activo',
        )
        cls.proyecto_febrero = Proyecto.objects.create(
            id='PROY-DATE-002',
            nombre='Proyecto Febrero',
            descripcion='Proyecto de febrero',
            fecha_inicio=date(2026, 2, 10),
            fecha_fin=date(2026, 2, 28),
            estado='activo',
        )

        cls.tarea_enero = TareaPorDesarrollar.objects.create(
            id='TASK-DATE-001',
            usuario=cls.developer,
            proyecto=cls.proyecto_enero,
            titulo='Tarea Enero',
            descripcion='Tarea de enero',
            fecha_vencimiento=date(2026, 1, 31),
            estado='pendiente',
        )
        cls.tarea_febrero = TareaPorDesarrollar.objects.create(
            id='TASK-DATE-002',
            usuario=cls.developer,
            proyecto=cls.proyecto_febrero,
            titulo='Tarea Febrero',
            descripcion='Tarea de febrero',
            fecha_vencimiento=date(2026, 2, 28),
            estado='pendiente',
        )

        TareaPorDesarrollar.objects.filter(
            pk=cls.tarea_enero.pk
        ).update(
            fecha_creacion=timezone.make_aware(
                datetime(2026, 1, 15, 10, 0)
            )
        )
        TareaPorDesarrollar.objects.filter(
            pk=cls.tarea_febrero.pk
        ).update(
            fecha_creacion=timezone.make_aware(
                datetime(2026, 2, 15, 10, 0)
            )
        )

        cls.tarea_enero.refresh_from_db()
        cls.tarea_febrero.refresh_from_db()

    def setUp(self):
        self.client = Client()
        self.assertTrue(
            self.client.login(
                username=self.admin.username,
                password='test-pass-123',
            )
        )

    def test_project_start_date_filters_results(self):
        response = self.client.get(
            reverse('buscar_proyectos'),
            {'fecha_inicio_desde': '2026-02-01'},
        )
        resultados = response.context['resultados']

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.proyecto_febrero, resultados)
        self.assertNotIn(self.proyecto_enero, resultados)

    def test_project_end_date_filters_results(self):
        response = self.client.get(
            reverse('buscar_proyectos'),
            {'fecha_inicio_hasta': '2026-01-31'},
        )
        resultados = response.context['resultados']

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.proyecto_enero, resultados)
        self.assertNotIn(self.proyecto_febrero, resultados)

    def test_invalid_project_date_returns_empty_without_error(self):
        response = self.client.get(
            reverse('buscar_proyectos'),
            {'fecha_inicio_desde': '2026-02-31'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            response.context['resultados'].exists()
        )

    def test_task_start_date_filters_results(self):
        response = self.client.get(
            reverse('buscar_tareas'),
            {'fecha_inicio_desde_tarea': '2026-02-01'},
        )
        resultados = response.context['resultados_tareas']

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.tarea_febrero, resultados)
        self.assertNotIn(self.tarea_enero, resultados)

    def test_task_end_date_filters_results(self):
        response = self.client.get(
            reverse('buscar_tareas'),
            {'fecha_inicio_hasta_tarea': '2026-01-31'},
        )
        resultados = response.context['resultados_tareas']

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.tarea_enero, resultados)
        self.assertNotIn(self.tarea_febrero, resultados)

    def test_invalid_task_date_returns_empty_without_error(self):
        response = self.client.get(
            reverse('buscar_tareas'),
            {'fecha_inicio_desde_tarea': 'fecha-invalida'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            response.context['resultados_tareas'].exists()
        )

    def test_project_excel_accepts_invalid_dates(self):
        response = self.client.get(
            reverse('exportar_proyectos_excel'),
            {
                'fecha_inicio_desde': '2026-02-31',
                'fecha_inicio_hasta': 'fecha-invalida',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_task_excel_accepts_invalid_dates(self):
        response = self.client.get(
            reverse('exportar_tareas_excel'),
            {
                'fecha_inicio_desde_tarea': '2026-02-31',
                'fecha_inicio_hasta_tarea': 'fecha-invalida',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_searches_and_exports_reuse_shared_filters(self):
        import inspect
        from .views import (
            buscar_proyectos,
            buscar_tareas,
            exportar_proyectos_excel,
            exportar_tareas_excel,
        )

        expected_helpers = (
            (buscar_proyectos, '_filtrar_proyectos('),
            (exportar_proyectos_excel, '_filtrar_proyectos('),
            (buscar_tareas, '_filtrar_tareas('),
            (exportar_tareas_excel, '_filtrar_tareas('),
        )

        for view, helper_name in expected_helpers:
            with self.subTest(view=view.__name__):
                self.assertIn(
                    helper_name,
                    inspect.getsource(view),
                )

    def test_project_pdf_works_without_previous_search(self):
        response = self.client.get(
            reverse('generar_informe_pdf_busqueda'),
            {'fecha_inicio_desde': '2026-02-01'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/pdf',
        )

    def test_task_pdf_works_without_previous_search(self):
        response = self.client.get(
            reverse('generar_informe_pdf_tareas'),
            {'fecha_inicio_desde_tarea': '2026-02-01'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/pdf',
        )

    def test_project_pdf_reuses_shared_filter(self):
        import inspect
        from .views import generar_informe_pdf_busqueda

        source = inspect.getsource(
            generar_informe_pdf_busqueda
        )

        self.assertIn('_filtrar_proyectos(', source)
        self.assertNotIn('request.session', source)

    def test_task_pdf_reuses_shared_filter(self):
        import inspect
        from .views import generar_informe_pdf_tareas

        source = inspect.getsource(
            generar_informe_pdf_tareas
        )

        self.assertIn('_filtrar_tareas(', source)
        self.assertNotIn('request.session', source)

    def test_project_and_task_searches_do_not_store_results_in_session(self):
        import inspect
        from .views import buscar_proyectos, buscar_tareas

        for view in (buscar_proyectos, buscar_tareas):
            with self.subTest(view=view.__name__):
                source = inspect.getsource(view)
                self.assertNotIn('request.session', source)


class UserSearchFilterTests(TestCase):
    """Pruebas del Hito 6: búsqueda y exportación filtrada de usuarios."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = Usuario.objects.create_superuser(
            username='admin_search',
            email='admin_search@test.local',
            password='test-pass-123',
            nombre='Administrador',
            apellido='Principal',
        )
        cls.developer = Usuario.objects.create_user(
            username='roberto_dev',
            email='roberto.gonzalez@test.local',
            password='test-pass-123',
            nombre='Roberto',
            apellido='González',
        )
        cls.other_developer = Usuario.objects.create_user(
            username='maria_code',
            email='maria.soto@test.local',
            password='test-pass-123',
            nombre='María',
            apellido='Soto',
        )
        from datetime import datetime
        from django.utils import timezone

        Usuario.objects.filter(pk=cls.admin.pk).update(
            fecha_creacion=timezone.make_aware(
                datetime(2026, 1, 5, 10, 0)
            )
        )
        Usuario.objects.filter(pk=cls.developer.pk).update(
            fecha_creacion=timezone.make_aware(
                datetime(2026, 1, 15, 10, 0)
            )
        )
        Usuario.objects.filter(pk=cls.other_developer.pk).update(
            fecha_creacion=timezone.make_aware(
                datetime(2026, 2, 10, 10, 0)
            )
        )

        cls.admin.refresh_from_db()
        cls.developer.refresh_from_db()
        cls.other_developer.refresh_from_db()

    def setUp(self):
        self.client = Client()

    def _login(self, user):
        self.assertTrue(
            self.client.login(
                username=user.username,
                password='test-pass-123',
            )
        )

    def _search(self, term):
        self._login(self.admin)
        return self.client.get(
            reverse('buscar_usuarios'),
            {'nombre_o_apellido': term},
        )

    def test_anonymous_user_is_redirected_to_login(self):
        response = self.client.get(reverse('buscar_usuarios'))

        self.assertEqual(response.status_code, 302)
        self.assertIn('/accounts/login/', response.url)

    def test_developer_cannot_access_global_user_search(self):
        self._login(self.developer)

        response = self.client.get(reverse('buscar_usuarios'))

        self.assertEqual(response.status_code, 302)

    def test_admin_can_access_global_user_search(self):
        self._login(self.admin)

        response = self.client.get(reverse('buscar_usuarios'))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'busqueda.html')

    def test_search_by_nombre(self):
        response = self._search('Roberto')
        usuarios = response.context['resultados_usuarios']

        self.assertIn(self.developer, usuarios)
        self.assertNotIn(self.other_developer, usuarios)

    def test_search_by_apellido(self):
        response = self._search('Soto')
        usuarios = response.context['resultados_usuarios']

        self.assertIn(self.other_developer, usuarios)
        self.assertNotIn(self.developer, usuarios)

    def test_search_by_username(self):
        response = self._search('roberto_dev')
        usuarios = response.context['resultados_usuarios']

        self.assertIn(self.developer, usuarios)
        self.assertNotIn(self.other_developer, usuarios)

    def test_search_by_email(self):
        response = self._search('maria.soto@test.local')
        usuarios = response.context['resultados_usuarios']

        self.assertIn(self.other_developer, usuarios)
        self.assertNotIn(self.developer, usuarios)

    def test_search_is_partial(self):
        response = self._search('bert')
        usuarios = response.context['resultados_usuarios']

        self.assertIn(self.developer, usuarios)

    def test_search_is_case_insensitive(self):
        response = self._search('rObErTo')
        usuarios = response.context['resultados_usuarios']

        self.assertIn(self.developer, usuarios)

    def test_empty_term_returns_all_users_without_error(self):
        response = self._search('')
        usuarios = response.context['resultados_usuarios']

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.admin, usuarios)
        self.assertIn(self.developer, usuarios)
        self.assertIn(self.other_developer, usuarios)

    def test_whitespace_term_returns_all_users_without_error(self):
        response = self._search('   ')
        usuarios = response.context['resultados_usuarios']

        self.assertEqual(response.status_code, 200)
        self.assertEqual(usuarios.count(), Usuario.objects.count())

    def test_search_without_matches_returns_empty_queryset(self):
        response = self._search('usuario-inexistente')
        usuarios = response.context['resultados_usuarios']

        self.assertEqual(response.status_code, 200)
        self.assertFalse(usuarios.exists())
        self.assertContains(
            response,
            'No se encontraron usuarios con los criterios ingresados.',
        )

    def test_search_preserves_filter_value_in_template(self):
        response = self._search('Roberto')

        self.assertContains(
            response,
            'value="Roberto"',
            html=False,
        )

    def test_excel_uses_real_user_fields_and_filter(self):
        from io import BytesIO
        import openpyxl

        self._login(self.admin)

        response = self.client.get(
            reverse('exportar_usuarios_excel'),
            {'nombre_o_apellido': 'Roberto'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = openpyxl.load_workbook(
            filename=BytesIO(response.content)
        )
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][1], self.developer.username)
        self.assertEqual(rows[1][2], self.developer.nombre)
        self.assertEqual(rows[1][3], self.developer.apellido)
        self.assertEqual(rows[1][4], self.developer.email)

    def test_valid_start_date_filters_users(self):
        self._login(self.admin)

        response = self.client.get(
            reverse('buscar_usuarios'),
            {'fecha_alta_desde': '2026-02-01'},
        )
        usuarios = response.context['resultados_usuarios']

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.other_developer, usuarios)
        self.assertNotIn(self.admin, usuarios)
        self.assertNotIn(self.developer, usuarios)

    def test_valid_end_date_filters_users(self):
        self._login(self.admin)

        response = self.client.get(
            reverse('buscar_usuarios'),
            {'fecha_alta_hasta': '2026-01-31'},
        )
        usuarios = response.context['resultados_usuarios']

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.admin, usuarios)
        self.assertIn(self.developer, usuarios)
        self.assertNotIn(self.other_developer, usuarios)

    def test_invalid_date_text_is_ignored(self):
        self._login(self.admin)

        response = self.client.get(
            reverse('buscar_usuarios'),
            {'fecha_alta_desde': 'fecha-invalida'},
        )
        usuarios = response.context['resultados_usuarios']

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            usuarios.count(),
            Usuario.objects.count(),
        )

    def test_impossible_date_is_ignored(self):
        self._login(self.admin)

        response = self.client.get(
            reverse('buscar_usuarios'),
            {'fecha_alta_desde': '2026-02-31'},
        )
        usuarios = response.context['resultados_usuarios']

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            usuarios.count(),
            Usuario.objects.count(),
        )

    def test_valid_date_is_applied_when_other_date_is_invalid(self):
        self._login(self.admin)

        response = self.client.get(
            reverse('buscar_usuarios'),
            {
                'fecha_alta_desde': '2026-02-01',
                'fecha_alta_hasta': 'fecha-invalida',
            },
        )
        usuarios = response.context['resultados_usuarios']

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.other_developer, usuarios)
        self.assertNotIn(self.admin, usuarios)
        self.assertNotIn(self.developer, usuarios)

    def test_invalid_dates_do_not_break_pdf_or_excel(self):
        self._login(self.admin)

        parameters = {
            'fecha_alta_desde': '2026-02-31',
            'fecha_alta_hasta': 'fecha-invalida',
        }

        pdf_response = self.client.get(
            reverse('generar_informe_pdf_usuarios'),
            parameters,
        )
        excel_response = self.client.get(
            reverse('exportar_usuarios_excel'),
            parameters,
        )

        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(
            pdf_response['Content-Type'],
            'application/pdf',
        )
        self.assertEqual(excel_response.status_code, 200)
        self.assertEqual(
            excel_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_pdf_and_excel_reuse_shared_filter(self):
        import inspect
        from .views import (
            buscar_usuarios,
            generar_informe_pdf_usuarios,
            exportar_usuarios_excel,
        )

        for view in (
            buscar_usuarios,
            generar_informe_pdf_usuarios,
            exportar_usuarios_excel,
        ):
            source = inspect.getsource(view)
            self.assertIn('_filtrar_usuarios(', source)


class UserDisplayNameTests(TestCase):
    """Pruebas del Hito 7: uso coherente de nombre y apellido."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = Usuario.objects.create_superuser(
            username='admin_names',
            email='admin_names@test.local',
            password='test-pass-123',
            nombre='Nombre Administrador',
            apellido='Apellido Administrador',
            first_name='LegacyAdmin',
            last_name='LegacySurname',
        )
        cls.developer = Usuario.objects.create_user(
            username='developer_names',
            email='developer_names@test.local',
            password='test-pass-123',
            nombre='Roberto',
            apellido='González',
            first_name='LegacyName',
            last_name='LegacyLastName',
        )

    def setUp(self):
        self.client = Client()

    def _login(self, user):
        self.assertTrue(
            self.client.login(
                username=user.username,
                password='test-pass-123',
            )
        )

    def test_views_and_templates_do_not_use_legacy_name_fields(self):
        from pathlib import Path

        app_path = Path(__file__).resolve().parent
        files = [
            app_path / 'views.py',
            app_path / 'templates' / 'busqueda.html',
            app_path / 'templates' / 'evaluaciones.html',
            app_path / 'templates' / 'home.html',
            app_path / 'templates' / 'proyectos.html',
            app_path / 'templates' / 'revisar_tareas.html',
            app_path / 'templates' / 'tareas.html',
            app_path / 'templates' / 'usuarios.html',
        ]

        for file_path in files:
            source = file_path.read_text(encoding='utf-8')

            self.assertNotIn(
                'first_name',
                source,
                f'{file_path.name} todavía usa first_name',
            )
            self.assertNotIn(
                'last_name',
                source,
                f'{file_path.name} todavía usa last_name',
            )

    def test_home_displays_custom_nombre(self):
        self._login(self.developer)

        response = self.client.get(reverse('home'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.developer.nombre)
        self.assertNotContains(response, self.developer.first_name)

    def test_users_page_displays_custom_name_and_surname(self):
        self._login(self.admin)

        response = self.client.get(reverse('usuarios'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'{self.developer.nombre} {self.developer.apellido}',
        )
        self.assertNotContains(
            response,
            f'{self.developer.first_name} {self.developer.last_name}',
        )

    def test_global_users_excel_uses_custom_name_fields(self):
        from io import BytesIO
        import openpyxl

        self._login(self.admin)

        response = self.client.get(
            reverse('exportar_todos_usuarios_excel')
        )

        self.assertEqual(response.status_code, 200)

        workbook = openpyxl.load_workbook(
            filename=BytesIO(response.content)
        )
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))

        developer_rows = [
            row for row in rows
            if row[1] == self.developer.username
        ]

        self.assertEqual(len(developer_rows), 1)

        row = developer_rows[0]
        self.assertEqual(row[2], self.developer.nombre)
        self.assertEqual(row[3], self.developer.apellido)
        self.assertNotEqual(row[2], self.developer.first_name)
        self.assertNotEqual(row[3], self.developer.last_name)


class ExcelFormulaInjectionTests(TestCase):
    """Pruebas de neutralización de fórmulas en exportaciones Excel."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = Usuario.objects.create_superuser(
            username='admin_excel_security',
            email='admin.excel.security@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Excel',
        )
        cls.dangerous_user = Usuario.objects.create_user(
            username='=SUM(A1:A2)',
            email='@correo.test',
            password='test-pass-123',
            nombre='+Nombre',
            apellido='-Apellido',
        )

    def setUp(self):
        self.client = Client()

    def _login(self):
        self.assertTrue(
            self.client.login(
                username=self.admin.username,
                password='test-pass-123',
            )
        )

    def test_excel_value_helper_neutralizes_formula_prefixes(self):
        from .views import _valor_excel_seguro

        dangerous_values = (
            '=SUM(A1:A2)',
            '+123',
            '-10+20',
            '@comando',
        )

        for value in dangerous_values:
            with self.subTest(value=value):
                self.assertEqual(
                    _valor_excel_seguro(value),
                    f"'{value}",
                )

    def test_excel_helpers_preserve_safe_values_and_types(self):
        from datetime import date, datetime

        from .views import (
            _fila_excel_segura,
            _valor_excel_seguro,
        )

        safe_values = (
            'Texto normal',
            '',
            42,
            3.5,
            None,
            date(2026, 8, 1),
            datetime(2026, 8, 1, 12, 30),
        )

        for value in safe_values:
            with self.subTest(value=value):
                self.assertEqual(
                    _valor_excel_seguro(value),
                    value,
                )

        self.assertEqual(
            _fila_excel_segura([
                '=FORMULA()',
                'Normal',
                10,
                None,
            ]),
            [
                "'=FORMULA()",
                'Normal',
                10,
                None,
            ],
        )

    def test_user_export_stores_dangerous_values_as_text(self):
        from io import BytesIO

        import openpyxl

        self._login()

        response = self.client.get(
            reverse('exportar_usuarios_excel')
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            (
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )

        workbook = openpyxl.load_workbook(
            filename=BytesIO(response.content),
            data_only=False,
        )
        worksheet = workbook.active

        self.assertEqual(
            [
                worksheet.cell(row=1, column=column).value
                for column in range(1, 7)
            ],
            [
                'ID',
                'Username',
                'Nombre',
                'Apellido',
                'Email',
                'Fecha de Registro',
            ],
        )

        user_row = None

        for row_number in range(2, worksheet.max_row + 1):
            if (
                worksheet.cell(row=row_number, column=1).value
                == self.dangerous_user.pk
            ):
                user_row = row_number
                break

        self.assertIsNotNone(user_row)

        expected_values = (
            "'=SUM(A1:A2)",
            "'+Nombre",
            "'-Apellido",
            "'@correo.test",
        )

        for column, expected_value in zip(
            range(2, 6),
            expected_values,
        ):
            with self.subTest(column=column):
                cell = worksheet.cell(
                    row=user_row,
                    column=column,
                )
                self.assertEqual(cell.value, expected_value)
                self.assertEqual(cell.data_type, 's')


class AuthenticationFlowCleanupTests(TestCase):
    """Pruebas del flujo seguro de cierre de sesión."""

    @classmethod
    def setUpTestData(cls):
        cls.user = Usuario.objects.create_user(
            username='authentication_flow_user',
            email='authentication.flow@test.local',
            password='test-pass-123',
            nombre='Authentication',
            apellido='Flow',
        )

    def setUp(self):
        self.client = Client()
        self.assertTrue(
            self.client.login(
                username=self.user.username,
                password='test-pass-123',
            )
        )

    def test_logout_get_is_rejected_and_preserves_session(self):
        response = self.client.get(reverse('logout'))

        self.assertEqual(response.status_code, 405)
        self.assertIn(
            '_auth_user_id',
            self.client.session,
        )

        home_response = self.client.get(reverse('home'))

        self.assertEqual(home_response.status_code, 200)

    def test_logout_post_closes_session(self):
        response = self.client.post(reverse('logout'))

        self.assertRedirects(
            response,
            reverse('login'),
            fetch_redirect_response=False,
        )
        self.assertNotIn(
            '_auth_user_id',
            self.client.session,
        )

    def test_base_template_uses_post_logout_form_with_csrf(self):
        response = self.client.get(reverse('home'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'action="{reverse("logout")}"',
            html=False,
        )
        self.assertContains(
            response,
            'method="post"',
            html=False,
        )
        self.assertContains(
            response,
            'name="csrfmiddlewaretoken"',
            html=False,
        )
        self.assertNotContains(
            response,
            f'href="{reverse("logout")}"',
            html=False,
        )

    def test_obsolete_custom_login_view_is_removed(self):
        from . import views

        self.assertFalse(
            hasattr(views, 'custom_login')
        )


class EvaluationConfigSingletonTests(TestCase):
    """Pruebas de configuración única para evaluaciones."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = Usuario.objects.create_superuser(
            username='admin_evaluation_config',
            email='admin.evaluation.config@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Configuración',
        )

    def setUp(self):
        self.client = Client()

    def _create_config(self, **overrides):
        from .models import EvaluacionConfig

        values = {
            'tiempo_entrega': 25,
            'complejidad_tarea': 25,
            'cumplimiento_requerimientos': 25,
            'calidad_codigo': 25,
            'nota_maxima': 100,
        }
        values.update(overrides)

        return EvaluacionConfig.objects.create(**values)

    def test_first_configuration_can_be_created_and_updated(self):
        from .models import EvaluacionConfig

        config = self._create_config()

        config.nota_maxima = 80
        config.save()
        config.refresh_from_db()

        self.assertEqual(config.nota_maxima, 80)
        self.assertEqual(
            EvaluacionConfig.objects.count(),
            1,
        )

    def test_second_configuration_is_rejected(self):
        from .models import EvaluacionConfig

        self._create_config()

        with self.assertRaisesRegex(
            ValueError,
            'Solo puede existir una configuración de evaluación.',
        ):
            self._create_config(nota_maxima=80)

        self.assertEqual(
            EvaluacionConfig.objects.count(),
            1,
        )

    def test_admin_disables_add_when_configuration_exists(self):
        from django.contrib import admin
        from django.test import RequestFactory

        from .models import EvaluacionConfig

        request = RequestFactory().get(
            '/admin/Reveloper/evaluacionconfig/'
        )
        request.user = self.admin

        model_admin = admin.site._registry[
            EvaluacionConfig
        ]

        self.assertTrue(
            model_admin.has_add_permission(request)
        )

        self._create_config()

        self.assertFalse(
            model_admin.has_add_permission(request)
        )

    def test_review_view_uses_default_without_configuration(self):
        self.assertTrue(
            self.client.login(
                username=self.admin.username,
                password='test-pass-123',
            )
        )

        response = self.client.get(
            reverse('revisar_tareas')
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context['nota_maxima'],
            100,
        )

    def test_review_view_uses_existing_configuration(self):
        self._create_config(nota_maxima=80)

        self.assertTrue(
            self.client.login(
                username=self.admin.username,
                password='test-pass-123',
            )
        )

        response = self.client.get(
            reverse('revisar_tareas')
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context['nota_maxima'],
            80,
        )


class MySQLStrictModeSettingsTests(TestCase):
    """Pruebas de configuración segura de MySQL."""

    def test_default_database_uses_mysql_strict_mode(self):
        from ProyectEspecial import settings as project_settings

        database = project_settings.DATABASES['default']
        options = database.get('OPTIONS', {})

        self.assertEqual(
            database['ENGINE'],
            'django.db.backends.mysql',
        )
        self.assertEqual(
            options.get('init_command'),
            "SET sql_mode='STRICT_TRANS_TABLES'",
        )

    def test_test_settings_continue_using_sqlite(self):
        from ProyectEspecial import test_settings

        database = test_settings.DATABASES['default']
        database_name = str(database['NAME'])

        self.assertEqual(
            database['ENGINE'],
            'django.db.backends.sqlite3',
        )
        self.assertTrue(
            database_name == ':memory:'
            or database_name.startswith('file:memorydb_')
        )
        self.assertNotIn(
            'init_command',
            database.get('OPTIONS', {}),
        )


class TemporaryReportFileCleanupTests(TestCase):
    """Pruebas de limpieza de recursos en informes gráficos PDF."""

    @classmethod
    def setUpTestData(cls):
        from datetime import date

        from .models import Evaluacion, Proyecto, Usuario

        cls.admin = Usuario.objects.create_superuser(
            username='admin_temp_pdf',
            email='admin_temp_pdf@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Temporal',
        )
        cls.developer = Usuario.objects.create_user(
            username='developer_temp_pdf',
            email='developer_temp_pdf@test.local',
            password='test-pass-123',
            nombre='Developer',
            apellido='Temporal',
        )
        cls.project = Proyecto.objects.create(
            id='PROJECT-TEMP-PDF',
            nombre='Proyecto temporal PDF',
            descripcion='Proyecto utilizado para probar informes.',
            fecha_inicio=date(2026, 1, 1),
            fecha_fin=date(2026, 12, 31),
            estado='activo',
        )
        cls.evaluation = Evaluacion.objects.create(
            titulo='Evaluación temporal',
            comentarios='Evaluación utilizada para generar el gráfico.',
            proyecto=cls.project,
            usuario=cls.developer,
            calificacion=85,
        )

    def setUp(self):
        from django.test import Client

        self.client = Client()

    def _login(self, user):
        self.assertTrue(
            self.client.login(
                username=user.username,
                password='test-pass-123',
            )
        )

    def _create_temp_path(self):
        import tempfile
        from pathlib import Path

        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix='.png',
        ) as temp_file:
            temp_path = temp_file.name

        self.addCleanup(
            lambda: Path(temp_path).unlink(missing_ok=True)
        )

        return temp_path

    def _patch_named_tempfile(self, temp_path):
        from unittest.mock import patch

        return patch(
            'Reveloper.views.tempfile.NamedTemporaryFile',
            side_effect=lambda *args, **kwargs: open(
                temp_path,
                'wb+',
            ),
        )

    def test_temporary_png_is_deleted_after_successful_pdf(self):
        from pathlib import Path

        self._login(self.developer)
        temp_path = self._create_temp_path()

        with self._patch_named_tempfile(temp_path):
            response = self.client.get(
                reverse('generar_informe_grafico_pdf')
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/pdf',
        )
        self.assertTrue(response.content.startswith(b'%PDF'))
        self.assertFalse(Path(temp_path).exists())

    def test_temporary_png_is_deleted_when_draw_image_fails(self):
        from pathlib import Path
        from unittest.mock import patch

        self._login(self.developer)
        temp_path = self._create_temp_path()

        with self._patch_named_tempfile(temp_path), patch(
            'Reveloper.views.canvas.Canvas.drawImage',
            side_effect=RuntimeError('drawImage falló'),
        ):
            with self.assertRaisesRegex(
                RuntimeError,
                'drawImage falló',
            ):
                self.client.get(
                    reverse('generar_informe_grafico_pdf')
                )

        self.assertFalse(Path(temp_path).exists())

    def test_admin_report_deletes_each_temporary_png(self):
        from pathlib import Path

        self._login(self.admin)
        temp_path = self._create_temp_path()

        with self._patch_named_tempfile(temp_path):
            response = self.client.get(
                reverse('generar_informe_grafico_pdf_admin')
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/pdf',
        )
        self.assertFalse(Path(temp_path).exists())

    def test_graph_resources_close_when_savefig_fails(self):
        from io import BytesIO
        from unittest.mock import patch

        from .views import generar_grafico_evaluaciones

        self._login(self.developer)
        request = type(
            'Request',
            (),
            {'user': self.developer},
        )()
        buffer = BytesIO()

        with patch(
            'Reveloper.views.BytesIO',
            return_value=buffer,
        ), patch(
            'Reveloper.views.plt.savefig',
            side_effect=RuntimeError('savefig falló'),
        ), patch(
            'Reveloper.views.plt.close',
        ) as close_mock:
            with self.assertRaisesRegex(
                RuntimeError,
                'savefig falló',
            ):
                generar_grafico_evaluaciones(
                    request,
                    self.developer,
                )

        close_mock.assert_called_once()
        self.assertTrue(buffer.closed)

    def test_graph_pdf_views_do_not_use_os_remove(self):
        import inspect

        from .views import (
            generar_informe_grafico_pdf_admin,
            generar_informe_grafico_pdf_desarrollador,
        )

        for view in (
            generar_informe_grafico_pdf_admin,
            generar_informe_grafico_pdf_desarrollador,
        ):
            with self.subTest(view=view.__name__):
                source = inspect.getsource(view)

                self.assertNotIn(
                    'os.remove(temp_file_path)',
                    source,
                )
                self.assertIn(
                    '_eliminar_archivo_temporal',
                    source,
                )
                self.assertIn('finally:', source)

    def test_graph_pdf_buffers_close_after_successful_response(self):
        from io import BytesIO
        from unittest.mock import patch

        for url_name, user in (
            ('generar_informe_grafico_pdf_admin', self.admin),
            ('generar_informe_grafico_pdf', self.developer),
        ):
            with self.subTest(url_name=url_name):
                self._login(user)

                pdf_buffer = BytesIO()
                graph_buffer = BytesIO()

                with patch(
                    'Reveloper.views.BytesIO',
                    side_effect=[
                        pdf_buffer,
                        graph_buffer,
                    ],
                ):
                    response = self.client.get(
                        reverse(url_name)
                    )

                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response['Content-Type'],
                    'application/pdf',
                )
                self.assertTrue(
                    response.content.startswith(b'%PDF')
                )
                self.assertTrue(pdf_buffer.closed)
                self.assertTrue(graph_buffer.closed)

    def test_graph_pdf_buffers_close_when_canvas_creation_fails(self):
        from io import BytesIO
        from unittest.mock import patch

        for url_name, user in (
            ('generar_informe_grafico_pdf_admin', self.admin),
            ('generar_informe_grafico_pdf', self.developer),
        ):
            with self.subTest(url_name=url_name):
                self._login(user)
                pdf_buffer = BytesIO()

                with patch(
                    'Reveloper.views.BytesIO',
                    return_value=pdf_buffer,
                ), patch(
                    'Reveloper.views.canvas.Canvas',
                    side_effect=RuntimeError(
                        'canvas falló'
                    ),
                ):
                    with self.assertRaisesRegex(
                        RuntimeError,
                        'canvas falló',
                    ):
                        self.client.get(
                            reverse(url_name)
                        )

                self.assertTrue(pdf_buffer.closed)


class PdfSpecialCharacterHandlingTests(TestCase):
    """Pruebas del escape de texto en informes PDF."""

    def test_pdf_paragraph_helper_escapes_reportlab_markup(self):
        from unittest.mock import Mock, patch

        from .views import _parrafo_pdf_seguro

        style = Mock()
        text = (
            'Proyecto A & Proyecto B '
            '<b>aparente</b> 5 > 3'
        )

        with patch(
            'Reveloper.views.Paragraph'
        ) as paragraph:
            result = _parrafo_pdf_seguro(
                text,
                style,
            )

        paragraph.assert_called_once_with(
            (
                'Proyecto A &amp; Proyecto B '
                '&lt;b&gt;aparente&lt;/b&gt; 5 &gt; 3'
            ),
            style,
        )
        self.assertIs(
            result,
            paragraph.return_value,
        )

    def test_pdf_paragraph_preserves_special_characters_as_text(self):
        from reportlab.lib.styles import getSampleStyleSheet

        from .views import _parrafo_pdf_seguro

        text = (
            'Error <crítico> & '
            'Texto <b>aparente</b> 5 > 3'
        )

        paragraph = _parrafo_pdf_seguro(
            text,
            getSampleStyleSheet()['Normal'],
        )

        self.assertEqual(
            paragraph.getPlainText(),
            text,
        )

    def test_project_pdf_accepts_special_characters(self):
        from datetime import date

        from .models import Proyecto, TareaPorDesarrollar

        admin = Usuario.objects.create_superuser(
            username='admin_pdf_special',
            email='admin.pdf.special@test.local',
            password='test-pass-123',
            nombre='Admin & PDF',
            apellido='<Especial>',
        )

        project = Proyecto.objects.create(
            id='PROY-PDF-SPECIAL',
            nombre='Proyecto A & Proyecto B',
            descripcion=(
                'Error <crítico> y '
                'texto <b>aparente</b> 5 > 3'
            ),
            fecha_inicio=date(2026, 8, 1),
            fecha_fin=date(2026, 8, 31),
            estado='activo',
        )

        TareaPorDesarrollar.objects.create(
            id='TASK-PDF-SPECIAL',
            titulo='Tarea <urgente> & revisión',
            descripcion=(
                'Comparación 5 > 3 '
                'y etiqueta <i>aparente</i>'
            ),
            fecha_vencimiento=date(2026, 8, 15),
            estado='pendiente',
            proyecto=project,
            usuario=admin,
        )

        self.client.force_login(admin)

        response = self.client.get(
            reverse('generate_pdf')
        )

        self.assertEqual(
            response.status_code,
            200,
        )
        self.assertEqual(
            response['Content-Type'],
            'application/pdf',
        )
        self.assertTrue(
            response.content.startswith(b'%PDF')
        )


class NPlusOneQueryOptimizationTests(TestCase):
    """Pruebas de consultas constantes en vistas con relaciones."""

    @classmethod
    def setUpTestData(cls):
        from datetime import date

        from .models import Proyecto, TareaPorDesarrollar

        cls.admin = Usuario.objects.create_superuser(
            username='admin_query_tests',
            email='admin.query.tests@test.local',
            password='test-pass-123',
            nombre='Admin',
            apellido='Consultas',
        )

        cls.developer = Usuario.objects.create_user(
            username='developer_query_tests',
            email='developer.query.tests@test.local',
            password='test-pass-123',
            nombre='Developer',
            apellido='Consultas',
        )

        cls.project = Proyecto.objects.create(
            id='PROJECT-QUERY-1',
            nombre='Proyecto de consultas',
            descripcion='Proyecto inicial',
            fecha_inicio=date(2026, 8, 1),
            fecha_fin=date(2026, 8, 31),
            estado='activo',
        )

        TareaPorDesarrollar.objects.create(
            id='TASK-QUERY-1',
            usuario=cls.developer,
            proyecto=cls.project,
            titulo='Tarea inicial',
            descripcion='Tarea para medir consultas',
            fecha_vencimiento=date(2026, 8, 15),
            estado='pendiente',
        )

    def setUp(self):
        self.client.force_login(self.admin)

    def _capture_request_queries(
        self,
        url_name,
        parameters=None,
    ):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        with CaptureQueriesContext(connection) as context:
            response = self.client.get(
                reverse(url_name),
                parameters or {},
            )

        self.assertEqual(response.status_code, 200)

        return len(context)

    def _create_additional_data(self):
        from datetime import date

        from .models import Proyecto, TareaPorDesarrollar

        for index in range(2, 7):
            developer = Usuario.objects.create_user(
                username=f'developer_query_{index}',
                email=f'developer.query.{index}@test.local',
                password='test-pass-123',
                nombre=f'Developer {index}',
                apellido='Consultas',
            )

            project = Proyecto.objects.create(
                id=f'PROJECT-QUERY-{index}',
                nombre=f'Proyecto {index}',
                descripcion='Proyecto adicional',
                fecha_inicio=date(2026, 8, 1),
                fecha_fin=date(2026, 8, 31),
                estado='activo',
            )

            TareaPorDesarrollar.objects.create(
                id=f'TASK-QUERY-{index}',
                usuario=developer,
                proyecto=project,
                titulo=f'Tarea {index}',
                descripcion='Tarea adicional',
                fecha_vencimiento=date(2026, 8, 15),
                estado='pendiente',
            )

    def test_dashboard_query_count_does_not_grow_with_users(self):
        initial_queries = self._capture_request_queries(
            'dashboard'
        )

        self._create_additional_data()

        expanded_queries = self._capture_request_queries(
            'dashboard'
        )

        self.assertEqual(
            expanded_queries,
            initial_queries,
        )

    def test_users_page_query_count_does_not_grow_with_users(self):
        initial_queries = self._capture_request_queries(
            'usuarios'
        )

        self._create_additional_data()

        expanded_queries = self._capture_request_queries(
            'usuarios'
        )

        self.assertEqual(
            expanded_queries,
            initial_queries,
        )

    def test_projects_page_query_count_does_not_grow_with_projects(self):
        initial_queries = self._capture_request_queries(
            'proyectos'
        )

        self._create_additional_data()

        expanded_queries = self._capture_request_queries(
            'proyectos'
        )

        self.assertEqual(
            expanded_queries,
            initial_queries,
        )

    def test_project_search_query_count_does_not_grow(self):
        parameters = {
            'fecha_inicio_desde': '2026-08-01',
        }

        initial_queries = self._capture_request_queries(
            'buscar_proyectos',
            parameters,
        )

        self._create_additional_data()

        expanded_queries = self._capture_request_queries(
            'buscar_proyectos',
            parameters,
        )

        self.assertEqual(
            expanded_queries,
            initial_queries,
        )

    def test_project_pdf_query_count_does_not_grow(self):
        initial_queries = self._capture_request_queries(
            'generate_pdf'
        )

        self._create_additional_data()

        expanded_queries = self._capture_request_queries(
            'generate_pdf'
        )

        self.assertEqual(
            expanded_queries,
            initial_queries,
        )

    def test_user_pdf_query_count_does_not_grow(self):
        initial_queries = self._capture_request_queries(
            'generate_user_pdf'
        )

        self._create_additional_data()

        expanded_queries = self._capture_request_queries(
            'generate_user_pdf'
        )

        self.assertEqual(
            expanded_queries,
            initial_queries,
        )

    def test_filtered_project_pdf_query_count_does_not_grow(self):
        parameters = {
            'fecha_inicio_desde': '2026-08-01',
        }

        initial_queries = self._capture_request_queries(
            'generar_informe_pdf_busqueda',
            parameters,
        )

        self._create_additional_data()

        expanded_queries = self._capture_request_queries(
            'generar_informe_pdf_busqueda',
            parameters,
        )

        self.assertEqual(
            expanded_queries,
            initial_queries,
        )

    def test_global_users_excel_query_count_does_not_grow(self):
        initial_queries = self._capture_request_queries(
            'exportar_todos_usuarios_excel'
        )

        self._create_additional_data()

        expanded_queries = self._capture_request_queries(
            'exportar_todos_usuarios_excel'
        )

        self.assertEqual(
            expanded_queries,
            initial_queries,
        )

    def test_admin_graph_pdf_query_count_does_not_grow(self):
        initial_queries = self._capture_request_queries(
            'generar_informe_grafico_pdf_admin'
        )

        self._create_additional_data()

        expanded_queries = self._capture_request_queries(
            'generar_informe_grafico_pdf_admin'
        )

        self.assertEqual(
            expanded_queries,
            initial_queries,
        )


class PdfResponseBufferCleanupTests(TestCase):
    """Pruebas del cierre de buffers en informes PDF."""

    def test_helper_closes_buffer_after_successful_build(self):
        from io import BytesIO
        from unittest.mock import Mock, patch

        from .views import _construir_respuesta_pdf

        buffer = BytesIO()
        document = Mock()

        def build_story(story):
            buffer.write(b'%PDF-test-content')

        document.build.side_effect = build_story

        with patch(
            'Reveloper.views.io.BytesIO',
            return_value=buffer,
        ), patch(
            'Reveloper.views.SimpleDocTemplate',
            return_value=document,
        ):
            response = _construir_respuesta_pdf(
                ['contenido de prueba']
            )

        document.build.assert_called_once_with(
            ['contenido de prueba']
        )
        self.assertEqual(
            response['Content-Type'],
            'application/pdf',
        )
        self.assertEqual(
            response.content,
            b'%PDF-test-content',
        )
        self.assertTrue(buffer.closed)

    def test_helper_closes_buffer_when_build_fails(self):
        from io import BytesIO
        from unittest.mock import Mock, patch

        from .views import _construir_respuesta_pdf

        buffer = BytesIO()
        document = Mock()
        document.build.side_effect = RuntimeError(
            'build falló'
        )

        with patch(
            'Reveloper.views.io.BytesIO',
            return_value=buffer,
        ), patch(
            'Reveloper.views.SimpleDocTemplate',
            return_value=document,
        ):
            with self.assertRaisesRegex(
                RuntimeError,
                'build falló',
            ):
                _construir_respuesta_pdf(
                    ['contenido de prueba']
                )

        self.assertTrue(buffer.closed)

    def test_helper_preserves_original_build_exception(self):
        from io import BytesIO
        from unittest.mock import Mock, patch

        from .views import _construir_respuesta_pdf

        original_error = ValueError(
            'error original del documento'
        )
        buffer = BytesIO()
        document = Mock()
        document.build.side_effect = original_error

        with patch(
            'Reveloper.views.io.BytesIO',
            return_value=buffer,
        ), patch(
            'Reveloper.views.SimpleDocTemplate',
            return_value=document,
        ):
            with self.assertRaises(ValueError) as context:
                _construir_respuesta_pdf([])

        self.assertIs(context.exception, original_error)
        self.assertTrue(buffer.closed)

    def test_project_pdf_still_returns_valid_pdf_response(self):
        self.client.force_login(
            Usuario.objects.create_superuser(
                username='admin_pdf_buffer',
                email='admin_pdf_buffer@test.local',
                password='test-pass-123',
                nombre='Admin',
                apellido='PDF',
            )
        )

        response = self.client.get(
            reverse('generate_pdf')
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/pdf',
        )
        self.assertTrue(
            response.content.startswith(b'%PDF')
        )

    def test_all_simple_pdf_views_use_shared_helper(self):
        import inspect

        from .views import (
            generate_evaluation_pdf,
            generate_pdf,
            generate_task_pdf,
            generate_user_pdf,
            generar_informe_pdf_busqueda,
            generar_informe_pdf_tareas,
            generar_informe_pdf_usuarios,
        )

        views = (
            generate_pdf,
            generate_task_pdf,
            generate_evaluation_pdf,
            generate_user_pdf,
            generar_informe_pdf_busqueda,
            generar_informe_pdf_tareas,
            generar_informe_pdf_usuarios,
        )

        for view in views:
            with self.subTest(view=view.__name__):
                source = inspect.getsource(view)

                self.assertIn(
                    'return _construir_respuesta_pdf(story)',
                    source,
                )
                self.assertNotIn(
                    'SimpleDocTemplate(',
                    source,
                )
                self.assertNotIn(
                    'buffer = io.BytesIO()',
                    source,
                )
                self.assertNotIn(
                    'doc.build(story)',
                    source,
                )